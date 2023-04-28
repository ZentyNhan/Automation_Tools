import os
import re
import openpyxl
import os.path
import ctypes

from openpyxl import styles
from openpyxl.styles import Font, Color
from openpyxl.worksheet.datavalidation import DataValidation

from os import system

DBCInfomation = {
                    'Node' : '',
                    'ID' : '',
                    'MSG' : '',
                    'MSGID' : '',
                    'Cyclic' : '',
                    'Signal' : '',
                    'Dir' : '',
                    'StartBit' : '',
                    'Length' : 0,
                    'Min' : 0.0,
                    'Max' : 0.0,
                    'Factor' : 0.0,
                    'Offset' : 0.0,
                    'InitValue' : 0.0,
                    'Unit' : '',
                    'Description' : '',
                    'DLC' : '',
                    'Endian' : '',
                    'Sign' : '',
                    'Flag' : '',
                }

InitInfomation = {
                    'Signal' : '',
                    'Value' : 0.0
                 }

LongSymbolInfo = {
                    'Signal' : '',
                    'Value' : ''
                 }

DescriptionInfo = {
                    'Signal' : '',
                    'Desc' : ''
                  }

MSGCycleInfo = {
                    'MSG' : '',
                    'Value' : ''
                }

MSGInfo = {
    		'Name' : '',
            'Cycle' : 0,
            'DLC' : 0,
            'Index' : 0,
            'Channel' : 0,
            'ID' : '',
          }

ColID = {
            'NodeCol' : 1,
            'MSGCol' : 2,
            'MSGIDCol' : 3, 
            'DLCCol' : 4,
            'DirCol' : 5,
            'CyclCol' : 6,
            'SigCol' : 7,
            'LengthCol' : 8, 
            'StartBitCol' : 9, 
            'EndianCol' : 10,
            'SignCol' : 11,
            'MinCol' : 12,
            'MaxCol' : 13,
            'FactCol' : 14,
            'OffsetCol' : 15,
            'InitCol' : 16,
            'UnitCol' : 17,
            'DescCol' : 18,
            # ReqIDCol' : 
            # NormCol' : 
            # ConvCol' : 
            'ALVCHKCol' : 19,
            'TSCol' : 20,
            'DataTypeCol' : 21 
        }

tLData = [
            'Node',
            'Message',
            'MessageID',
            'DLC',
            'DIR',
            'Cylic',
            'Signal',
            'Length',
            'StartBit(MSB)',
            'Endian',
            'Sign',
            'MinValue',
            'MaxValue',
            'Factor',
            'Offset',
            'InitValue',
            'Unit',
            'Description',
            # 'ReqID
            # 'Norm
            # 'ConvFact
            'ALV/CHK_Flag'
        ]

Flag = {
        'Flag_Sheet_Pub' : 0,
        'Flag_Sheet_Pri' : 0,
        'Flag_BO' : False,
        'Flag_Compare' : True,
        'Flag_Fill' : False
        }

InitValue = {
                't_fill' : '',
                't_sort' : '',
            }

FSI_Data = "ALV, CHK, VehicleSpeed, VehicleStandstill, EngineStatus, Gear, Seatbelt, DriverDoorStatus, TurnIndicators, SteeringAngle, Steering_AngleSpeed, EPBStatus, EPBSwitch, MasCylBrakePressure, MasCylBrake_PressureValid, AccPedalPos, AccPedal_PosValid, WheelSpeed"
CLARA_Data = "ALV, CHK, ACCMode, TauGapSet, TargetSpeed, TargetAccelerationforACC, TargetAccelerationforAEB_Enable, TargetAccelerationforAEB_Value, ACCDecelerationToStopRequest, DriveOffRequest, ACCTargetObject, WheelSpeed_Status, WheelDirection, WheelSpeed_FR, WheelSpeed_FL, WheelSpeed_RR, WheelSpeed_RL, WheelPulse_FR, WheelPulse_FL, WheelPulse_RR, WheelPulse_RL, VehicleStandstill, MasterCylinderPressure, Yawrate, VehicleSpeed, LongitudinalAcceleration, LateralAcceleration, TurnIndicator, SteeringWheelAngle, SteeringWheelSpeed, DriverOverrideSignal, AccelerationPedalPosition, BrakePedalStatus, EngineRunning, EngineSpeed, ACCMainSwitch, ACCTipUp, ACCTipDown, ACCOff, ACCResume, ACCSet, ACCSetPlusTime, ACCSetMinusTime, Odometer, GearShiftLevel, KeyStatus"



t_DBCInfo = dict(DBCInfomation)
InitInfo = []
t_InitInfo = dict(InitInfomation)
Signals = []
t_Signals = dict(DBCInfomation)
DescInfos = []
t_DescInfos = dict(DescriptionInfo)
CycleInfo = []
t_CycleInfo = dict(MSGCycleInfo)
LongSymbols = []
t_LongSymbols = dict(LongSymbolInfo)

# Macro MessageBox
MessageBox = ctypes.windll.user32.MessageBoxW
IDOK = 0
IDCANCEL = 2
IDABORT = 3
IDYES = 6
IDNO = 7

def Replace_f(strInput, Pattern, strRplace):
    if re.search(Pattern, strInput) != None : 
        return re.sub(Pattern, strRplace, re.search(Pattern, strInput).group())
    else:
        return strInput

def Search_f(strInput, Pattern):
    if strInput != None:
        if re.search(Pattern, strInput) != None : 
            return re.search(Pattern, strInput).group()
        else:
            return ""
    else:
        return ""

def Format_FSI(i, sheet):
    data_val = DataValidation(type="list",formula1='"{}"'.format(FSI_Data))
    sheet.add_data_validation(data_val)
    data_val.add(sheet.cell(i, ColID['ALVCHKCol']))

def Format_CLARA(i, sheet):
    data_val = DataValidation(type="list",formula1='"{}"'.format(CLARA_Data))
    sheet.add_data_validation(data_val)
    data_val.add(sheet.cell(i, ColID['ALVCHKCol']))

def GetDBCInfo(DBCPath, NodeName):
    DBCInfo = {}
    # MSGInfo = {}
    DBC_Compare = {}
    try:
        DBC = open(DBCPath, "r", errors="ignore")
        txtDBC = DBC.readlines()
    except IOError:
        print("Can not open file.dbc")
    finally:
        DBC.close()

    for tMach in txtDBC:
        BO_Info = re.search('^BO_ .*', tMach)
        MSGK1 = ''
        if BO_Info != None:
            Patern = "BO_ (\d+) (\w+): (\d+) (\w+)"
            tNode = Replace_f(tMach, Patern, r"\4")
            tmsg = Replace_f(tMach, Patern, r"\2")
            tMSGID = Replace_f(tMach, Patern, r"\1")
            tDLC = Replace_f(tMach, Patern, r"\3")
            Flag['Flag_BO'] = True
            if tmsg != MSGK1:
                DBCInfo[tmsg] = dict()
                MSGK1 = tmsg
        
         
        Signal_Info = re.search('^[ ]SG_ .+', tMach)
        if Signal_Info != None:
            t_DBCInfo = dict(DBCInfomation)
            Patern = "SG_ (\w+).*? : (\w+)\|(\w+)@(.+?) \((.+?),(.+?)\) \[(.+?)\|(.+?)\].*?\"(.*)\"(.+)"
            temp = Replace_f(tMach, Patern, r"\1|\2|\3|\4|\5|\6|\7|\8|\9|\10")
            tResult = temp.split('|')
    
            if InitValue['t_sort'] == tmsg.strip():
                DBC_Compare[tmsg.strip()][tResult[0].strip()] = {} #len(DBCInfo)
            else:
                DBC_Compare[tmsg.strip()] = {}
                DBC_Compare[tmsg.strip()][tResult[0].strip()] = {} #len(DBCInfo)

            InitValue['t_sort'] = tmsg.strip()
            t_DBCInfo['MSG'] = tmsg.strip()
            t_DBCInfo['MSGID'] = tMSGID.strip()
            t_DBCInfo['DLC'] = tDLC.strip()
            t_DBCInfo['Signal'] = tResult[0].strip()
            t_DBCInfo['Length'] = tResult[2]
            # 0: motorola(big endian) tResult[1] is MSB, 1: intel(little endian) tResult[1] is LSB
            if "1" in tResult[3] != 0: 
                tStart = int(tResult[1].strip())
                tStart = (tStart + int(t_DBCInfo['Length']) - 1) #convert LSB to MSB
                t_DBCInfo['Endian'] = "Little"
            elif "0" in tResult[3] != 0:
                tStart = int(tResult[1].strip())
                t_DBCInfo['Endian'] = "Big"
            
            # -: Signed, +: Unsigned
            if "-" in tResult[3] != 0: 
                t_DBCInfo['Sign'] = "signed"
            elif "+" in tResult[3] != 0:
                t_DBCInfo['Sign'] = "unsigned"

            t_DBCInfo['StartBit'] = tStart
            t_DBCInfo['Factor'] = tResult[4]
            t_DBCInfo['Offset'] = tResult[5]
            t_DBCInfo['Min'] = tResult[6]
            t_DBCInfo['Max'] = tResult[7]
            t_DBCInfo['Unit'] = tResult[8]

            if (tNode == NodeName) or (NodeName in tResult[9]):
                t_DBCInfo['Node'] = NodeName
            else:
                t_DBCInfo['Node'] = tResult[9]
            
            if tNode == NodeName:
                t_DBCInfo['Dir'] = "RX"
            else:
                t_DBCInfo['Dir'] = "TX"

            tSig = {}
            tSig[t_DBCInfo['Signal']] = t_DBCInfo
            if t_DBCInfo['Node'] == NodeName:
                DBCInfo[tmsg].update(tSig)
    
    # Find init value
    for tMach in txtDBC:
        Patern = "^BA_ \"GenSigStartValue\" *?SG_ \d+ (\w+) (\d+)"
        tStr = re.search(Patern, tMach)
        if tStr != None:
            t_InitInfo = dict(InitInfomation)
            t_InitInfo['Signal'] = Replace_f(tMach, Patern, r"\1")
            t_InitInfo['Value'] = Replace_f(tMach, Patern, r"\2")
            InitInfo.append(t_InitInfo)

    # Find cyclic
    for tMach in txtDBC:
        Patern = "^BA_ \"GenMsgCycleTime\" *?BO_ (\d+) (\d+)"
        tStr = re.search(Patern, tMach)
        if tStr != None:
            t_CycleInfo = dict(MSGCycleInfo)
            t_CycleInfo['MSG'] = Replace_f(tMach, Patern, r"\1")
            t_CycleInfo['Value'] = Replace_f(tMach, Patern, r"\2")
            CycleInfo.append(t_CycleInfo)

    #  Find long symbol
    for tMach in txtDBC:
        Patern = "^BA_ \"SystemSignalLongSymbol\" *?SG_ (\d+) (\w+) \"(\w+)\""
        tStr = re.search(Patern, tMach)
        if tStr != None:
            t_LongSymbols = dict(LongSymbolInfo)
            t_LongSymbols['Signal'] = Replace_f(tMach, Patern, r"\2")
            t_LongSymbols['Value'] = Replace_f(tMach, Patern, r"\3")
            LongSymbols.append(t_LongSymbols)

    # Find Description
    for tMach in txtDBC:
        Patern = '^VAL_ \d+ (\w+)(.+?);'
        tStr = re.search(Patern, tMach)
        if tStr != None:
            t_DescInfos = dict(DescriptionInfo)
            t_DescInfos['Signal'] = Replace_f(tMach, Patern, r"\1")
            t_Des = Replace_f(tMach, Patern, r"\2")
            t_Des = re.findall('(\d+ \".+?)\"', t_Des)
            t_DescInfos['Desc'] = ''
            for temp in t_Des:
                t_DescInfos['Desc'] = t_DescInfos['Desc'].strip() + temp.replace(' "', '->').strip() + '| '
            DescInfos.append(t_DescInfos)
    
    # rectify signal 
    for tSignals in DBCInfo.values():
        for t_SignalInfor in tSignals.values():
            t_SignalInfor['InitValue'] = 0
            # Match initvalue
            for t_InitInfo in InitInfo:
                if t_InitInfo['Signal'] == t_SignalInfor['Signal']:
                    t_SignalInfor['InitValue'] = int(t_InitInfo['Value']) * float(t_SignalInfor['Factor']) + float(t_SignalInfor['Offset'])
                    break
            # Match Desc
            for t_DescInfos in DescInfos:
                if t_DescInfos['Signal'] == t_SignalInfor['Signal']:
                    t_SignalInfor['Description'] = t_DescInfos['Desc']
                    break
            # zip(sorted(InitInfo.keys()), sorted(DescInfos.keys())):
            # Match Cyclic
            for t_CycleInfo in CycleInfo:
                if t_CycleInfo['MSG'] == t_SignalInfor['MSGID']:
                    t_SignalInfor['Cyclic'] = t_CycleInfo['Value']
                    break
            # Match LongSymbol
            for t_LongSymbols in LongSymbols:
                if t_LongSymbols['Signal'] == t_SignalInfor['Signal']:
                    t_SignalInfor['Signal'] = t_LongSymbols['Value']
                    break        
   
    return DBCInfo

def GENtoExcel(MainFilePath, DBCInfo, Channel, type):   
    book = openpyxl.load_workbook(MainFilePath)
    sheet = book.active
    
    if Channel == "1":
        tWBName = "Output_dbc"       
    else:
        tWBName = "Output_pri_dbc"
     
    # Sheet exist in excel file
    for isheet in book.worksheets:
        if isheet.title == tWBName:
            sheet = book[isheet.title]
            if MessageBox(None,"Do you want to clean old Sheet","WARNING",0x40 | 0x3) == IDYES :
                EndRow = sheet.max_row
                EndCol = sheet.max_column           
                sheet.delete_rows(idx=1, amount=EndRow)
                sheet.delete_cols(idx=1, amount=EndCol)
            Flag['Flag_Sheet_Pub'] = 1
            break
    # Sheet not exist in excel file
    if Flag['Flag_Sheet_Pub'] == 0:
        tRename = MessageBox(None, "Do you want to re-name Worksheet: \"" + book.active.title + "\" to \"" + tWBName,'WARNING', 0x40 | 0x3)
        if tRename == IDYES:
            sheet = book.active
            sheet.title = tWBName
            book.save(MainFilePath)
            if MessageBox(None,"Do you want to clean old Sheet","WARNING",0x40 | 0x3) == IDYES:
                EndRow = sheet.max_row
                EndCol = sheet.max_column           
                sheet.delete_rows(idx=1, amount=EndRow)
                sheet.delete_cols(idx=1, amount=EndCol)
        elif tRename == IDNO :
            book.create_sheet(tWBName)
            sheet = book[tWBName]
        else:
            MessageBox(None,"Can't find worksheet" + tWBName,"WARNING", 0x40 | 0x4000)
            return
    
    for i in range(ColID['ALVCHKCol']):
        sheet.cell(row = 1, column = i +1).value = tLData[i]
    
    tAllNode = MessageBox(None, 'Do you want to get all Node?','WARNING', 0x40 | 0x3)
    fill_index = 2
    # for t_DBC_Info in DBCInfo:
    for tSignals in DBCInfo.values():
        for t_DBC_Info in tSignals.values():
            # t = ((NodeName.lower() in t_DBC_Info['Node'].lower()) and t_DBC_Info['Signal'] != '')
            if (tAllNode == IDNO and t_DBC_Info['Signal'] != ''):
                sheet.cell(fill_index, ColID['NodeCol']).value     = str(t_DBC_Info['Node'])
                sheet.cell(fill_index, ColID['MSGCol']).value      = str(t_DBC_Info['MSG'])
                sheet.cell(fill_index, ColID['DLCCol']).value      = str(t_DBC_Info['DLC'])
                sheet.cell(fill_index, ColID['DirCol']).value      = str(t_DBC_Info['Dir'])
                sheet.cell(fill_index, ColID['CyclCol']).value     = str(t_DBC_Info['Cyclic'])
                sheet.cell(fill_index, ColID['SigCol']).value      = str(t_DBC_Info['Signal'])
                sheet.cell(fill_index, ColID['LengthCol']).value   = str(t_DBC_Info['Length'])
                sheet.cell(fill_index, ColID['StartBitCol']).value = str(t_DBC_Info['StartBit'])
                sheet.cell(fill_index, ColID['MinCol']).value      = str(t_DBC_Info['Min'])
                sheet.cell(fill_index, ColID['MaxCol']).value      = str(t_DBC_Info['Max'])
                sheet.cell(fill_index, ColID['FactCol']).value     = str(t_DBC_Info['Factor'])
                sheet.cell(fill_index, ColID['OffsetCol']).value   = str(t_DBC_Info['Offset'])
                sheet.cell(fill_index, ColID['InitCol']).value     = str(round(t_DBC_Info['InitValue'], 2))
                sheet.cell(fill_index, ColID['DescCol']).value     = str(t_DBC_Info['Description'])
                sheet.cell(fill_index, ColID['UnitCol']).value     = str(t_DBC_Info['Unit'])
                sheet.cell(fill_index, ColID['EndianCol']).value   = str(t_DBC_Info['Endian'])
                sheet.cell(fill_index, ColID['SignCol']).value     = str(t_DBC_Info['Sign'])
                sheet.cell(fill_index, ColID['MSGIDCol']).value    = str(hex(int(t_DBC_Info['MSGID']))).upper()
                
                fill_1 = styles.PatternFill(fill_type='solid', start_color='BBD3DB', end_color='BBD3DB')
                fill_2 = styles.PatternFill(fill_type='solid', start_color='C7DF7B', end_color='C7DF7B')

                if fill_index == 2:
                    Flag['Flag_Fill'] =  True
                elif InitValue['t_fill'] != t_DBC_Info['MSG']:
                    Flag['Flag_Fill'] = not Flag['Flag_Fill']

                if Flag['Flag_Fill']:
                    for j in range(1, ColID['ALVCHKCol'] + 1):  #+1
                        sheet.cell(fill_index,j).fill = fill_2
                else:
                    for j in range(1, ColID['ALVCHKCol'] + 1):  #+1
                        sheet.cell(fill_index, j).fill = fill_1
                InitValue['t_fill'] = t_DBC_Info['MSG']

                if type == "FSI":
                    Format_FSI(fill_index, sheet)
                elif type == "CLARA":
                    Format_CLARA(fill_index, sheet)

                fill_index += 1
    
    book.save(MainFilePath)


if __name__ == "__main__":
    MainFilePath = r'\\hc-ut40966c\D\UHC1HC\Gen5\Tool\fsi_gen5_toolchain\Database\DBC.xlsx'
    DBCPath = r'\\hc-ut40966c\D\UHC1HC\Gen5\Tool\fsi_gen5_toolchain\Database\MRRACAN_A20_A29_Combined.dbc'
    DBCPriPath = r'\\hc-ut40966c\D\UHC1HC\Gen5\Tool\fsi_gen5_toolchain\Database\GAC_RFC_VFC_DASy_V5.1_20200328.dbc'
    NodeNamePub = 'MRR'
    NodeNamePri = 'RadarFrontCenter'
    Channel = '1'
    type = 'FSI'
    Data_DBC = GetDBCInfo(DBCPath, NodeNamePub)
    GENtoExcel(MainFilePath, Data_DBC, Channel, type)