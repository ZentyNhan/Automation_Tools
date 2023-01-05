'''
################################### Infomation ###################################
# Tool Name    : IP Tool 
# Author       : Nguyen Thanh Nhan
# Function     : This tool support to create "import Participants" excel file
# Verion       : V1.0
# Release note : none
# Release date : Dec 2nd 2023
##################################################################################
'''
from xml.dom.expatbuilder import ElementInfo
from sqlite3 import connect
from PyQt5 import QtGui, QtWidgets
from PyQt5.QtGui import QStandardItemModel
from PyQt5.QtGui import QStandardItem
from PyQt5.QtWidgets import QMainWindow, QApplication , QDialog, QTabWidget,QDateEdit, QHBoxLayout, QGridLayout, QGroupBox, QVBoxLayout, QWidget, QLabel, QLineEdit, QDialogButtonBox, QMessageBox, QPushButton, QFileDialog, QTextEdit, QListWidget, QAbstractItemView, QComboBox
from PyQt5.QtCore import QThread, pyqtSignal, QObject, QDate, QTime, QDateTime
from PyQt5.QtWidgets import QCalendarWidget
from PyQt5.uic import loadUi
import sys
import datetime
import os, webbrowser
from faker import Faker
import random
from PyQt5 import QtTest
from PyQt5 import QtCore
from PyQt5.QtCore import QThread
from PyQt5.QtCore import QTimer
from Interface import Ui_GenerateTool
from PyQt5.QtWidgets import QDateEdit
from termcolor import colored 
from datetime import datetime
from datetime import date


#Excel Processing:
import openpyxl
from openpyxl import workbook,worksheet 
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font,PatternFill,Border,Side

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_GenerateTool()
        self.ui.setupUi(self)

        #Variables:
        #ANCHOR - Paths
        self.Direction_path = os.path.dirname(os.path.realpath(__file__))
        self.ImportTemplate_path = os.path.join(self.Direction_path, "Template","Import", "participant_course_import_template.xlsx")

        #ANCHOR - Configuration:
        #Current Date:
        self.Current_date = datetime.now()
        self.Global_date = self.Current_date.strftime("%d-%m-%Y_%H.%M.%S")
        #Report:
        self.WB_Report = openpyxl.load_workbook(filename=self.ImportTemplate_path, read_only=False, keep_vba=False,data_only=False, keep_links=True)
        
        #Add Worksheet name:
        self.WS_Report = self.WB_Report['Members']
        #Create Faker object:
        self.fake = Faker()
        self.Member_Gender = ['Female', 'Male']

        #ANCHOR - Infomation:
        self.Member_Quanlity = 1 #Adjust number of participants you need to test

        ##-----Events-----##:
        self.ui.Generate_pushButton.clicked.connect(self.Gen_IP_file)

        

    #ANCHOR - Functions:
    def Gen_IP_file(self):
        if self.ui.NumberOfParticipant_spinBox.text() == '0':
            self.ui.Notification_lineEdit.setStyleSheet("color: rgb(255, 0, 0);")
            self.ui.Notification_lineEdit.setText("Invalid number.Please correct number of participants")
        else:
            self.Member_Quanlity = int(self.ui.NumberOfParticipant_spinBox.text())
            for Num in range(2,self.Member_Quanlity+2):
                self.WS_Report['A{0}'.format(Num)] = self.fake.name()
                self.WS_Report['B{0}'.format(Num)] = random.choice(self.Member_Gender)
                self.WS_Report['C{0}'.format(Num)] = '=DATE({0},{1},{2})'.format( 
                                                                    str(random.randint(2000,2022)).zfill(1),  
                                                                    str(random.randint(1,12)).zfill(2),
                                                                    str(random.randint(1,30)).zfill(2)
                )
                self.WS_Report['D{0}'.format(Num)] = '028{0}'.format(str(random.randint(1,9999999)).zfill(7))
                self.WS_Report['E{0}'.format(Num)] = '09{0}'.format(str(random.randint(1,99999999)).zfill(8))
                self.WS_Report['F{0}'.format(Num)] = 'Test_{0}@gmail.com'.format(str(random.randint(1,9999999999)).zfill(10))
                self.WS_Report['G{0}'.format(Num)] = 'married'
                self.WS_Report['H{0}'.format(Num)] = 'DH'
                self.WS_Report['I{0}'.format(Num)] = int('231{0}'.format(str(random.randint(1,999999)).zfill(6)))
                self.WS_Report['J{0}'.format(Num)] = int('{0}{1}'.format(str(random.randint(1,9)).zfill(1), str(random.randint(0,999999999)).zfill(9)))
                self.WS_Report['K{0}'.format(Num)] = '=DATE({0},{1},{2})'.format( 
                                                                    str(random.randint(2000,2022)).zfill(1),  
                                                                    str(random.randint(1,12)).zfill(2),
                                                                    str(random.randint(1,30)).zfill(2)
                )
                self.WS_Report['L{0}'.format(Num)] = 'TP. Hà Nội'
                self.WS_Report['M{0}'.format(Num)] = 'TP. Hà Nội'
                self.WS_Report['N{0}'.format(Num)] = 'No'
                self.WS_Report['O{0}'.format(Num)] = 'VietNam'
                self.WS_Report['P{0}'.format(Num)] = 'Nam AD'
                self.WS_Report['Q{0}'.format(Num)] = 'AD1239'
                self.WS_Report['R{0}'.format(Num)] = ''
                self.WS_Report['S{0}'.format(Num)] = 'FWO'
                self.WS_Report['T{0}'.format(Num)] = '=DATE({0},{1},{2})'.format( 
                                                                    str(random.randint(2000,2022)).zfill(1),  
                                                                    str(random.randint(1,12)).zfill(2),
                                                                    str(random.randint(1,30)).zfill(2)
                )                                                                
                self.WS_Report['U{0}'.format(Num)] = 'Eric'

            #Save file:
            self.WB_Report.save('{0}/GEN/participant_course_import_{1}.xlsx'.format(self.Direction_path,self.Global_date))
            print('\n\n|-------------------- Genarated successfully --------------------|\n\n')
            
            #-----------------------------------#
            Processing_Satus = 'Processing' #Local variable
            for i in range(10):        
                self.ui.Notification_lineEdit.setStyleSheet("color: rgb(0, 0, 0);") #set text color
                Processing_Satus += '.'
                self.ui.Notification_lineEdit.setText(Processing_Satus)
                QtTest.QTest.qWait(100)
            self.ui.Notification_lineEdit.setStyleSheet("font-weight: bold; color: green") #set text color
            self.ui.Notification_lineEdit.setText('Genarated Successfully!')
            #-----------------------------------#


if __name__ == "__main__":
    app = QApplication(sys.argv) 
    main_win = MainWindow()
    main_win.show()
    sys.exit(app.exec())