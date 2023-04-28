from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import datetime
import collections
import re 
import os
import numpy as np
import collections

#path file EDA Employee
EDA_Employee = load_workbook("BGSV-EDA-Employee_List.xlsx")
EDA_Employee_ws = EDA_Employee['Associate-01.06.2022']

#path file Project List
Project_List = load_workbook("Projects_list_2022_Radar_Running_Overview_RBVH_edit.xlsx")
Project_List_ws = Project_List['Q1_2022']

def Employee_Dict(EDA_Employee_ws):
	EmployeeDict = {}
	Employee_No = ""
	Associate_Name = ""
	Group = ""
	for i in range(1, EDA_Employee_ws.max_column + 1):
		char = get_column_letter(i)
		if EDA_Employee_ws[char + str(1)].value == "Employee No.":
			Employee_No = char
		if EDA_Employee_ws[char + str(1)].value == "Associate Name":
			Associate_Name = char
		if EDA_Employee_ws[char + str(1)].value == "Group":
			Group = char
	for i in range(1, EDA_Employee_ws.max_row + 1):
		if EDA_Employee_ws[Group + str(i)].value == "MS/EDA3-XC":
			EmployeeDict_temp = {}
			EmployeeDict_temp['Name'] = str(EDA_Employee_ws[Associate_Name + str(i)].value)
			EmployeeDict_temp['NumberID'] = str(EDA_Employee_ws[Employee_No + str(i)].value)
			EmployeeDict[str(EDA_Employee_ws[Associate_Name + str(i)].value)] = EmployeeDict_temp
	return EmployeeDict

def Employee_Dict_Number(EDA_Employee_ws):
	EmployeeDict = {}
	Employee_No = ""
	Associate_Name = ""
	Group = ""
	num = 0
	for i in range(1, EDA_Employee_ws.max_column + 1):
		char = get_column_letter(i)
		if EDA_Employee_ws[char + str(1)].value == "Employee No.":
			Employee_No = char
		if EDA_Employee_ws[char + str(1)].value == "Associate Name":
			Associate_Name = char
		if EDA_Employee_ws[char + str(1)].value == "Group":
			Group = char
	for i in range(1, EDA_Employee_ws.max_row + 1):
		if EDA_Employee_ws[Group + str(i)].value == "MS/EDA3-XC":
			EmployeeDict_temp = {}
			EmployeeDict_temp['Name'] = str(EDA_Employee_ws[Associate_Name + str(i)].value)
			EmployeeDict_temp['NumberID'] = str(EDA_Employee_ws[Employee_No + str(i)].value)
			num += 1
			EmployeeDict[num] = EmployeeDict_temp
	return EmployeeDict

def Project_Dict(Project_List_ws):
	ProjectDict = {}
	Customer_Name = "A"
	Project_Name = "B"
	BM_Number = "C"
	Task_ID = "D"
	Project_Status  = "H"
	Status_Project = ["Running","Pending","Cancel","Cancelled","SOPed"]

	for i in range(1, Project_List_ws.max_row + 1):
		if Project_List_ws[Project_Status + str(i)].value in Status_Project :
			str_1 = str(Project_List_ws[Project_Name + str(i)].value).replace("+", " ").replace("/", " ").replace("_", " ").replace("-", " ").replace("\n", " ").replace("  ", " ").upper()
			ProjectDict_temp = {}
			ProjectDict_temp['CustomerName'] = str(Project_List_ws[Customer_Name + str(i)].value)
			ProjectDict_temp['ProjectName'] = str_1
			ProjectDict_temp['BM_Number'] = str(Project_List_ws[BM_Number + str(i)].value)
			ProjectDict_temp['Task_ID'] = str(Project_List_ws[Task_ID + str(i)].value)
			ProjectDict[str_1] = ProjectDict_temp
	return ProjectDict
    
EmployeeDict = Employee_Dict(EDA_Employee_ws)
EmployeeDict_Num = Employee_Dict_Number(EDA_Employee_ws)
ProjectDict = Project_Dict(Project_List_ws)



