########## SECTION: Library  ##########
#Others
import os
import sys
import PyQt5
import json
import time
import shutil
import datetime
import threading
import subprocess
import tkinter
import pandas
from tkinter import * 
from tkinter import messagebox
from bs4 import BeautifulSoup
from pyunpack import Archive
import zipfile

from datetime import datetime
from datetime import date
# from pytest import PytestCollectionWarning
from sqlite3 import connect
from PyQt5 import QtGui, QtWidgets, QtCore
from PyQt5.QtGui import QStandardItemModel
from PyQt5.QtGui import QStandardItem, QIntValidator
from PyQt5.QtWidgets import QMainWindow, QApplication , QDialog, QTabWidget,QDateEdit, QHBoxLayout, QGridLayout, QGroupBox, QVBoxLayout, QWidget, QLabel, QLineEdit, QDialogButtonBox, QMessageBox, QPushButton, QFileDialog, QTextEdit, QListWidget, QAbstractItemView, QComboBox
from PyQt5.QtCore import QThread, pyqtSignal, QObject, QDate, QTime, QDateTime
from PyQt5.QtGui import QIntValidator
from PyQt5.uic import loadUi
from PyQt5 import QtTest
from PyQt5.QtCore import QTimer

#From UI:
import SRPartner
from SRPartner import Ui_SRP_MainWindow

#From Lib:
from lib.Func import Func_INS
from lib.config import Config_INS
from lib.ui.Configuration_ui import *
from lib.can_engine import * #Diag
from lib.diagnostic import * #Diag

#Excel Processing:
from openpyxl import Workbook 
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Color, Fill
from openpyxl.styles import Alignment, Font,PatternFill,Border,Side
from openpyxl.utils.exceptions import InvalidFileException
########## !SECTION: Library  ##########

########## SECTION: Communicate  ##########
class Communicate(QThread):
    progress = pyqtSignal(int, str, int)
    def __init__(self):
        super().__init__()
########## !SECTION: Communicate  ##########

class Help(QDialog, Configuration_ui):
    def __init__(self, parent=None):
        super(Help, self).__init__(parent)
        self.setupUi(self)
        
########## SECTION: MainWindow  ##########
class SRPartner(QMainWindow):
    def __init__(self):
        super().__init__()
        self.Ter = Ui_SRP_MainWindow()
        self.Ter.setupUi(self)
        self.setWindowIcon(QtGui.QIcon('icon/icon.png'))
        self.statusBar().showMessage('Authored by HNG4HC')
        self.setWindowTitle('SRPartner')
        # allow only integers
        onlyInt = QIntValidator()
        onlyInt.setRange(0, 99)
        self.Ter.BL_lineEdit.setValidator(onlyInt)
        # self.Ter.RCV_lineEdit.setValidator(onlyInt)
        # Threading and Progress
        self.threads = []
        self.COM = Communicate()
        self.Ter.progressBar.setValue(0)
        # Translate:
        self.translate = QtCore.QCoreApplication.translate

        # ==================== #ANCHOR: Init variables: ====================================================================================================================================================================================================================================================================
        # ------------------------#ANCHOR: Combobox ------------------------:
        #Add items:
        self.Ter.PRJ_comboBox.addItems(Config_INS.Prj_Items)
        self.Ter.Type_comboBox.addItems(Config_INS.Type_Items)
        self.Ter.RCV_comboBox.addItems(Config_INS.RCV_Items)
        self.Ter.H2A_comboBox.addItems(Config_INS.H2A_Items)
        self.Ter.Pos_comboBox.addItems(Config_INS.Pos_Items)
        self.Ter.Tester_comboBox.addItems(Config_INS.Tester_Items)
        self.Ter.SWPCM_comboBox.addItems(Config_INS.SWPCM_Items)
        self.Ter.TC_comboBox.addItems(Config_INS.TC_Items)
        
        # Get Current path:
        self.Current_path = os.path.dirname(__file__)

        # Tab:
        Tab = {
            'Info'        : 0,
            'SR Check'    : 1,
            'FSI'         : 2,
            'HexToAscii'  : 3
        }
        
        # Checkbox:
        self.is_DS_check       = 'True'
        self.is_PS_check       = 'True'
        self.is_ES_check       = 'True'
        self.is_PhysDiag_check = 'True'
        self.is_FuncDiag_check = 'False'
        self.is_Sub01_check    = 'True'
        self.is_Sub02_check    = 'True'
        self.is_Sub03_check    = 'True'
        self.is_SPRMIB_check   = 'True'
        self.is_ActSes_F186    = 'True'
        self.CheckList_Name    = ''
        self.Protocol_Name     = ''

        # USER/Database path:
        USER              = os.getlogin()
        self.temp         = os.path.join(r'C:\Users', USER.lower(), r'AppData\Local\Temp')
        self.DB_org       = os.path.join(self.Current_path, r'Template\Tooldatabase.json')
        self.DB_dirname   = os.path.join(r'C:\Users', USER.lower(), r'AppData\Local\Temp\SRPartnerTool')
        self.Tooldatabase = os.path.join(self.DB_dirname, 'Tooldatabase.json')

        # Create database in temp:
        SRPartner.createDatabase(self)

        #Check TesterPresent:
        SRPartner.Run_testerpresentChangedAction_Func(self)

        
        #Set/Load init value:
        self.LoadInit_path = self.Tooldatabase
        with open(self.LoadInit_path, 'r') as LoadInit_json:
            self.LoadInit_dict = json.load(LoadInit_json)
        self.LoadJSON()

        # ==================== #ANCHOR: Events: ====================================================================================================================================================================================================================================================================
        # ------------------------#ANCHOR: TEST ------------------------:
        self.Ter.TEST_Open_Button.clicked.connect(self.TEST)
        # ------------------------#ANCHOR: Tab widget ------------------------:
        self.Ter.Info_tabWidget.currentChanged.connect(self.Run_TabChanged_Func)
        # ------------------------#ANCHOR: Buttton ------------------------:
        #ANCHOR - Tab Info:
        # CONTROL:
        self.Ter.GD_Button.clicked.connect(self.Run_GenDocument_Func)
        self.Ter.Ticket_Open_Button.clicked.connect(self.Run_TicketOpen_Func)
        self.Ter.LRF_Open_Button.clicked.connect(self.Run_LRFOpen_Func)
        self.Ter.TL_Open_Button.clicked.connect(self.Run_TLOpen_Func)
        self.Ter.SW_Open_Button.clicked.connect(self.Run_SWOpen_Func)
        self.Ter.WF_Open_Button.clicked.connect(self.Run_WFOpen_Func)
        self.Ter.DB_Open_Button.clicked.connect(self.Run_DBOpen_Func)
        self.Ter.CSW_Button.clicked.connect(self.Run_SWCopy_Func)
        self.Ter.GBR_Button.clicked.connect(self.Run_GenReportBackup_Func)
        self.Ter.GSR_Button.clicked.connect(self.Run_GenReportSuzhou_Func)
        self.Ter.GSRCR_Button.clicked.connect(self.Run_GenReportSRCheck_Func)
        self.Ter.WFP_Browse_Button.clicked.connect(self.Run_WFP_Browse_Func)
        self.Ter.SWP_Browse_Button.clicked.connect(self.Run_SWP_Browse_Func)
        self.Ter.WFP_Clear_Button.clicked.connect(self.Run_WFPTextClear_Func)
        self.Ter.SWP_Clear_Button.clicked.connect(self.Run_SWPTextClear_Func)
        self.Ter.BKS_Clear_Button.clicked.connect(self.Run_BKSTextClear_Func)
        self.Ter.SS_Clear_Button.clicked.connect(self.Run_SSTextClear_Func)
        self.Ter.PN_Clear_Button.clicked.connect(self.Run_PNTextClear_Func)
        self.Ter.SRCR_Clear_Button.clicked.connect(self.Run_SRCRTextClear_Func)
        # Org Link:
        self.Ter.SS_Open_Button.clicked.connect(self.Run_SSOpen_Func)
        self.Ter.SRCR_Open_Button.clicked.connect(self.Run_SRCROpen_Func)
        self.Ter.SWBuild_Open_Button.clicked.connect(self.Run_SWBuildOpen_Func)
        self.Ter.BKS_Open_Button.clicked.connect(self.Run_BKSOpen_Func)
        # Report Server:
        self.Ter.Report_BKS_Open_Button.clicked.connect(self.Run_ReportBKSOpen_Func)
        self.Ter.Report_SRCR_Open_Button.clicked.connect(self.Run_ReportSRCROpen_Func)
        self.Ter.Report_SS_Open_Button.clicked.connect(self.Run_ReportSSOpen_Func)
        self.Ter.Report_Protocol_Open_Button.clicked.connect(self.Run_ReportProtocolOpen_Func)
        #ANCHOR - Tab SR-Check:
        # Filepath:
        self.Ter.SB_Open_Button.clicked.connect(self.Run_SBOpen_Func)
        self.Ter.Diatester_Open_Button.clicked.connect(self.Run_DiatesterOpen_Func)
        self.Ter.CANalyzer_Open_Button.clicked.connect(self.Run_CANalyzerOpen_Func)
        self.Ter.CANape_Open_Button.clicked.connect(self.Run_CANapeOpen_Func)
        self.Ter.FaultID_Open_Button.clicked.connect(self.Run_FaultIDOpen_Func)
        self.Ter.PubDBC_Open_Button.clicked.connect(self.Run_PubDBCOpen_Func)
        self.Ter.PriDBC_Open_Button.clicked.connect(self.Run_PriDBCOpen_Func)
        self.Ter.Oxflash_Open_Button.clicked.connect(self.Run_OxflashOpen_Func)
        self.Ter.Nxflash_Open_Button.clicked.connect(self.Run_NxflashOpen_Func)
        self.Ter.ESBBOX_Open_Button.clicked.connect(self.Run_ESBBOXOpen_Func)
        self.Ter.STIL_Open_Button.clicked.connect(self.Run_STILOpen_Func)
        self.Ter.SB_toolButton.clicked.connect(self.Run_SB_Browse_Func)
        self.Ter.Diatester_toolButton.clicked.connect(self.Run_Diatester_Browse_Func)
        self.Ter.CANalyzer_toolButton.clicked.connect(self.Run_CANalyzer_Browse_Func)
        self.Ter.CANape_toolButton.clicked.connect(self.Run_CANape_Browse_Func)
        self.Ter.FaultID_toolButton.clicked.connect(self.Run_FaultID_Browse_Func)
        self.Ter.PubDBC_toolButton.clicked.connect(self.Run_PubDBC_Browse_Func)
        self.Ter.PriDBC_toolButton.clicked.connect(self.Run_PriDBC_Browse_Func)
        self.Ter.Oxflash_toolButton.clicked.connect(self.Run_Oxflash_Browse_Func)
        self.Ter.Nxflash_toolButton.clicked.connect(self.Run_Nxflash_Browse_Func)
        self.Ter.ESBBOX_toolButton.clicked.connect(self.Run_ESBBOX_Browse_Func)
        self.Ter.STIL_toolButton.clicked.connect(self.Run_STIL_Browse_Func)
        self.Ter.UDS_toolButton.clicked.connect(self.Run_UDS_Browse_Func)
        self.Ter.Case_1_pushButton.clicked.connect(self.Run_Case_1_Func)
        self.Ter.Case_3_pushButton.clicked.connect(self.Run_Case_3_Func)
        self.Ter.ResetECU_pushButton.clicked.connect(self.Run_ResetECU_Func)
        # Control:
        self.Ter.Send_diag_pushButton.clicked.connect(self.Run_SendDiagnostic_Func)
        self.Ter.ESPON_Open_Button.clicked.connect(self.Run_ESBBox_on_Func)
        self.Ter.ESPOFF_Open_Button.clicked.connect(self.Run_ESBBox_off_Func)
        
        #ANCHOR - Tab FSI:
        self.Ter.FSI_SB_Open_Button.clicked.connect(self.Run_FSI_SBOpen_Func)
        self.Ter.FSI_Diatester_Open_Button.clicked.connect(self.Run_FSI_DiatesterOpen_Func)
        self.Ter.FSI_CANalyzer_Open_Button.clicked.connect(self.Run_FSI_CANalyzerOpen_Func)
        self.Ter.FSI_CANape_Open_Button.clicked.connect(self.Run_FSI_CANapeOpen_Func)
        self.Ter.FSI_FaultID_Open_Button.clicked.connect(self.Run_FSI_FaultIDOpen_Func)
        self.Ter.FSI_PubDBC_Open_Button.clicked.connect(self.Run_FSI_PubDBCOpen_Func)
        self.Ter.FSI_PriDBC_Open_Button.clicked.connect(self.Run_FSI_PriDBCOpen_Func)
        self.Ter.FSI_SB_toolButton.clicked.connect(self.Run_FSI_SB_Browse_Func)
        self.Ter.FSI_Diatester_toolButton.clicked.connect(self.Run_FSI_Diatester_Browse_Func)
        self.Ter.FSI_CANalyzer_toolButton.clicked.connect(self.Run_FSI_CANalyzer_Browse_Func)
        self.Ter.FSI_CANape_toolButton.clicked.connect(self.Run_FSI_CANape_Browse_Func)
        self.Ter.FSI_FaultID_toolButton.clicked.connect(self.Run_FSI_FaultID_Browse_Func)
        self.Ter.FSI_PubDBC_toolButton.clicked.connect(self.Run_FSI_PubDBC_Browse_Func)
        self.Ter.FSI_PriDBC_toolButton.clicked.connect(self.Run_FSI_PriDBC_Browse_Func)
        self.Ter.Flash_Open_Button.clicked.connect(self.Run_OpenFlashCANape_Func)
        self.Ter.FSI_Ticket_Open_Button.clicked.connect(self.Run_TicketOpen_Func)
        self.Ter.ClaraTE_Open_Button.clicked.connect(self.Run_ClaraTEOpen_Func)
        #ANCHOR - Tab HexToAscii:
        self.Ter.H2A_pushButton.clicked.connect(self.Run_HexToAscii_Func)
        # ------------------------#ANCHOR: Combobox ------------------------:
        self.Ter.PRJ_comboBox.currentTextChanged.connect(self.Run_ChooseVariant_Func)
        self.Ter.H2A_comboBox.currentTextChanged.connect(self.Run_ChooseConvertMode_Func)
        self.Ter.Pos_comboBox.currentTextChanged.connect(self.Run_ChoosePosition_Func)
        self.Ter.RCV_comboBox.currentTextChanged.connect(self.Run_ChooseRCV_Func)
        self.Ter.Var_comboBox.currentTextChanged.connect(self.Run_ChooseVar_Func)
        self.Ter.Type_comboBox.currentTextChanged.connect(self.Run_ChooseType_Func)
        self.Ter.Tester_comboBox.currentTextChanged.connect(self.Run_ChooseTester_Func)
        self.Ter.SWPCM_comboBox.currentTextChanged.connect(self.Run_ChooseSWPCM_Func)
        self.Ter.TC_comboBox.currentTextChanged.connect(self.Run_ChooseTC_Func)
        # ------------------------#ANCHOR: ProgessBar ------------------------:
        self.COM.progress.connect(self.Progress)
        # ------------------------#ANCHOR: Tool bar ------------------------:
        self.Ter.actionAbout_SRPartner.triggered.connect(self.actionAbout)
        self.Ter.actionConfiguration.triggered.connect(self.actionConfiguration)
        # ------------------------#ANCHOR: CheckBox ------------------------:
        #ANCHOR - Tab SRCheck:
        self.Ter.TP_checkBox.stateChanged.connect(self.Run_testerpresentChangedAction_Func)
        self.Ter.AS_checkBox.stateChanged.connect(self.Run_ActSesCheck_Func)
        self.Ter.DS_checkBox.stateChanged.connect(self.Run_DefaultSessionCheck_Func)
        self.Ter.PS_checkBox.stateChanged.connect(self.Run_ProgrammingSessionCheck_Func)
        self.Ter.ES_checkBox.stateChanged.connect(self.Run_ExtendedSessionCheck_Func)
        self.Ter.PhysDiag_checkBox.stateChanged.connect(self.Run_PhysDiagCheck_Func)
        self.Ter.FuncDiag_checkBox.stateChanged.connect(self.Run_FuncDiagCheck_Func)
        self.Ter.Sub01_checkBox.stateChanged.connect(self.Run_Sub01Check_Func)
        self.Ter.Sub02_checkBox.stateChanged.connect(self.Run_Sub02Check_Func)
        self.Ter.Sub03_checkBox.stateChanged.connect(self.Run_Sub03Check_Func)
        self.Ter.SPRMIB_checkBox.stateChanged.connect(self.Run_SPRMIBCheck_Func)
        # ------------------------#ANCHOR: LineEdit ------------------------:
        #ANCHOR - Tab Info:
        self.Ter.WFP_lineEdit.textChanged.connect(self.Run_WFP_lineEditChanged_Func)
        self.Ter.SWP_lineEdit.textChanged.connect(self.Run_SWP_lineEditChanged_Func)
        self.Ter.PN_lineEdit.textChanged.connect(self.Run_PN_lineEditChanged_Func)
        self.Ter.BL_lineEdit.textChanged.connect(self.Run_BL_lineEditChanged_Func)
        self.Ter.RCV_lineEdit.textChanged.connect(self.Run_RCV_lineEditChanged_Func)
        self.Ter.HWPCM_lineEdit.textChanged.connect(self.Run_HWPCM_lineEditChanged_Func)
        self.Ter.TPM_lineEdit.textChanged.connect(self.Run_TPM_lineEditChanged_Func)
        self.Ter.BKS_lineEdit.textChanged.connect(self.Run_BKS_lineEditChanged_Func)
        self.Ter.SS_lineEdit.textChanged.connect(self.Run_SS_lineEditChanged_Func)
        self.Ter.SRCR_lineEdit.textChanged.connect(self.Run_SRCR_lineEditChanged_Func)
        self.Ter.TTicket_lineEdit.textChanged.connect(self.Run_TTicket_lineEditChanged_Func)
        self.Ter.Reviewer_lineEdit.textChanged.connect(self.Run_Reviewer_lineEditChanged_Func)
        self.Ter.TM_lineEdit.textChanged.connect(self.Run_TM_lineEditChanged_Func)
        # Hidden report:
        self.Ter.Report_BKS_lineEdit.textChanged.connect(self.Run_Report_BKS_lineEditChanged_Func)
        self.Ter.Report_SRCR_lineEdit.textChanged.connect(self.Run_Report_SRCR_lineEditChanged_Func)
        self.Ter.Report_SS_lineEdit.textChanged.connect(self.Run_Report_SS_lineEditChanged_Func)
        self.Ter.Report_Protocol_lineEdit.textChanged.connect(self.Run_Report_Protocol_lineEditChanged_Func)
        self.Ter.SWVer_lineEdit.textChanged.connect(self.Run_SWVer_lineEditChanged_Func)
        #ANCHOR - Tab SR-Check:
        self.Ter.SB_lineEdit.textChanged.connect(self.Run_SB_lineEditChanged_Func)
        self.Ter.Diatester_lineEdit.textChanged.connect(self.Run_Diatester_lineEditChanged_Func)
        self.Ter.CANalyzer_lineEdit.textChanged.connect(self.Run_CANalyzer_lineEditChanged_Func)
        self.Ter.CANape_lineEdit.textChanged.connect(self.Run_CANape_lineEditChanged_Func)
        self.Ter.FaultID_lineEdit.textChanged.connect(self.Run_FaultID_lineEditChanged_Func)
        self.Ter.PubDBC_lineEdit.textChanged.connect(self.Run_PubDBC_lineEditChanged_Func)
        self.Ter.PriDBC_lineEdit.textChanged.connect(self.Run_PriDBC_lineEditChanged_Func)
        self.Ter.Oxflash_lineEdit.textChanged.connect(self.Run_Oxflash_lineEditChanged_Func)
        self.Ter.Nxflash_lineEdit.textChanged.connect(self.Run_Nxflash_lineEditChanged_Func)
        self.Ter.ESBBOX_lineEdit.textChanged.connect(self.Run_ESBBOX_lineEditChanged_Func)
        self.Ter.STIL_lineEdit.textChanged.connect(self.Run_STIL_lineEditChanged_Func)
        # Control:
        self.Ter.TXID_lineEdit.textChanged.connect(self.Run_TXID_lineEditChanged_Func)
        self.Ter.RXID_lineEdit.textChanged.connect(self.Run_RXID_lineEditChanged_Func)
        self.Ter.AS_lineEdit.textChanged.connect(self.Run_AS_lineEditChanged_Func)
        self.Ter.UDS_lineEdit.textChanged.connect(self.Run_UDS_lineEditChanged_Func)
        self.Ter.UDSSheet_lineEdit.textChanged.connect(self.Run_UDSSheet_lineEditChanged_Func)
        
        #ANCHOR - Tab FSI-Check:
        self.Ter.FSI_SB_lineEdit.textChanged.connect(self.Run_FSI_SB_lineEditChanged_Func)
        self.Ter.FSI_Diatester_lineEdit.textChanged.connect(self.Run_Diatester_lineEditChanged_Func)
        self.Ter.FSI_CANalyzer_lineEdit.textChanged.connect(self.Run_CANalyzer_lineEditChanged_Func)
        self.Ter.FSI_CANape_lineEdit.textChanged.connect(self.Run_CANape_lineEditChanged_Func)
        self.Ter.FSI_FaultID_lineEdit.textChanged.connect(self.Run_FaultID_lineEditChanged_Func)
        self.Ter.FSI_PubDBC_lineEdit.textChanged.connect(self.Run_PubDBC_lineEditChanged_Func)
        self.Ter.FSI_PriDBC_lineEdit.textChanged.connect(self.Run_PriDBC_lineEditChanged_Func)

    # ==================== #ANCHOR: Functions: ====================================================================================================================================================================================================================================================================
    # -------------------- #ANCHOR: Checkbox Management --------------------:
    def Run_testerpresentChangedAction_Func(self):
        TXID = self.Ter.TXID_lineEdit.text()
        RXID = self.Ter.RXID_lineEdit.text()
        ISCANFD = True
        #Check CAN FD or not:
        if self.Ter.CANFD_checkBox.isChecked(): ISCANFD = True
        else:                                   ISCANFD = False
        #Config:
        txID, rxID, is_fd, VectorApp = Func_INS.Getconfiguration(TXID,RXID, ISCANFD)
        if txID != None:
            if (self.Ter.TP_checkBox.isChecked()):
                testerON(txID, rxID, is_fd, VectorApp)
            else:
                testerOff()
                
    def Run_ActSesCheck_Func(self):
        if self.Ter.AS_checkBox.isChecked():
            self.is_ActSes_F186 = True
            self.Ter.AS_lineEdit.setEnabled(False)
        else:
            self.is_ActSes_F186 = False
            self.Ter.AS_lineEdit.setEnabled(True)
        SRPartner.UpdateJSON(self)
                
    def Run_DefaultSessionCheck_Func(self):
        if self.Ter.DS_checkBox.isChecked():
            self.is_DS_check = True
        else:
            self.is_DS_check = False
        SRPartner.UpdateJSON(self)
            
    def Run_ProgrammingSessionCheck_Func(self):
        if self.Ter.PS_checkBox.isChecked():
            self.is_PS_check = True
        else:
            self.is_PS_check = False
        SRPartner.UpdateJSON(self)
            
    def Run_ExtendedSessionCheck_Func(self):
        if self.Ter.ES_checkBox.isChecked():
            self.is_ES_check = True
        else:
            self.is_ES_check = False
        SRPartner.UpdateJSON(self)
            
    def Run_PhysDiagCheck_Func(self):
        if self.Ter.PhysDiag_checkBox.isChecked():
            self.is_PhysDiag_check = True
        else:
            self.is_PhysDiag_check = False
        SRPartner.UpdateJSON(self)
            
    def Run_FuncDiagCheck_Func(self):
        if self.Ter.FuncDiag_checkBox.isChecked():
            self.is_FuncDiag_check = True
        else:
            self.is_FuncDiag_check = False
        SRPartner.UpdateJSON(self)
            
    def Run_Sub01Check_Func(self):
        if self.Ter.Sub01_checkBox.isChecked():
            self.is_Sub01_check = True
        else:
            self.is_Sub01_check = False
        SRPartner.UpdateJSON(self)
            
    def Run_Sub02Check_Func(self):
        if self.Ter.Sub02_checkBox.isChecked():
            self.is_Sub02_check = True
        else:
            self.is_Sub02_check = False
        SRPartner.UpdateJSON(self)
            
    def Run_Sub03Check_Func(self):
        if self.Ter.Sub03_checkBox.isChecked():
            self.is_Sub03_check = True
        else:
            self.is_Sub03_check = False
        SRPartner.UpdateJSON(self)
            
    def Run_SPRMIBCheck_Func(self):
        if self.Ter.SPRMIB_checkBox.isChecked():
            self.is_SPRMIB_check = True
        else:
            self.is_SPRMIB_check = False
        SRPartner.UpdateJSON(self)


    # -------------------- #ANCHOR: Browse Management --------------------:
    def Run_WFP_Browse_Func(self):
        OpenObject = QFileDialog.getExistingDirectory(parent= self, caption="Open working folder", directory= self.Current_path)
        self.Ter.WFP_lineEdit.setText(OpenObject)

    def Run_SWP_Browse_Func(self):
        path = r'\\abtvdfs2.de.bosch.com\ismdfs\loc\szh\DA\Driving\SW_TOOL_Release'
        OpenObject = QFileDialog.getExistingDirectory(parent= self, caption="Open SW path", directory= path)
        self.Ter.SWP_lineEdit.setText(OpenObject)

    def Run_SB_Browse_Func(self):
        path = str(self.LoadInit_dict['TabSRCheck']['SR_Filepath']['SB_lineEdit']).replace('/','\\')
        if path == '':
            path = r'D:\.'
        OpenObject = QFileDialog.getExistingDirectory(parent= self, caption="Open sandbox folder", directory= path)
        if OpenObject == '':
            self.Ter.SB_lineEdit.setText(self.Ter.SB_lineEdit.text())
        else:
            self.Ter.SB_lineEdit.setText(OpenObject)
        SRPartner.UpdateJSON(self)
        
    def Run_Diatester_Browse_Func(self): 
        Func_INS.Tool_Browse_Func(self.Ter.SB_lineEdit, self.Ter.Diatester_lineEdit, 'All (*.cna*)')
        SRPartner.UpdateJSON(self)
        
    def Run_CANalyzer_Browse_Func(self):
        Func_INS.Tool_Browse_Func(self.Ter.SB_lineEdit, self.Ter.CANalyzer_lineEdit, 'All (*.cfg*)')
        SRPartner.UpdateJSON(self)
        
    def Run_CANape_Browse_Func(self):
        Func_INS.Tool_Browse_Func(self.Ter.SB_lineEdit, self.Ter.CANape_lineEdit, 'All (*.cna*)')
        SRPartner.UpdateJSON(self)
        
    def Run_FaultID_Browse_Func(self):
        Func_INS.Tool_Browse_Func(self.Ter.SB_lineEdit, self.Ter.FaultID_lineEdit, 'All (*.h*)')
        SRPartner.UpdateJSON(self)
        
    def Run_PubDBC_Browse_Func(self):
        Func_INS.Tool_Browse_Func(self.Ter.SB_lineEdit, self.Ter.PubDBC_lineEdit, 'All (*.dbc*)')
        SRPartner.UpdateJSON(self)
        
    def Run_PriDBC_Browse_Func(self):
        Func_INS.Tool_Browse_Func(self.Ter.SB_lineEdit, self.Ter.PriDBC_lineEdit, 'All (*.dbc*)')
        SRPartner.UpdateJSON(self)
        
    def Run_Case_1_Func(self):
        try:
            dt_string = datetime.now().strftime("[%d-%m-%Y_%H.%M.%S]")
            path          = self.Ter.UDS_lineEdit.text()
            sheet         = self.Ter.UDSSheet_lineEdit.text()
            DID_cols      = 'A'
            DIDName_col   = 'B'
            DIDLength_col = 'C'
            REPORT_TYPE   = 'html'
            TESTCASE      = '1.3 / 1.4'
            TXID  = int(self.Ter.TXID_lineEdit.text(),16)
            RXID  = int(self.Ter.RXID_lineEdit.text(),16)
            ISCANFD = True
            #Check CAN FD or not:
            if self.Ter.CANFD_checkBox.isChecked(): ISCANFD = True
            else:                                   ISCANFD = False
            #Config:
            if path == '':
                SRPartner.diagnostic_Console(self, 'Error', 'Error UDS path.Please check')
            else:
                #Result:
                RESULT = Func_INS.ReadDID_Func(path, sheet, DID_cols, DIDName_col, DIDLength_col, TXID, RXID, ISCANFD)
                #report:
                if REPORT_TYPE in 'Excel':
                    #Excel:
                    df = pandas.DataFrame(RESULT).T
                    try:    df.to_excel(f"{os.path.join(self.Ter.WFP_lineEdit.text(),'PIC', 'log', 'Test_log')}/Case_1_3_and_1_4_{dt_string}.xlsx")
                    except: df.to_excel(f"{self.Current_path}/temp/report/Case_1_3_and_1_4_{dt_string}.xlsx")
                else:
                    #HTML:
                    SRPartner.HTML_Report(self,RESULT,TESTCASE)
                SRPartner.diagnostic_Console(self, 'Success', f'Case {TESTCASE} done!')
        except Exception as e:
            SRPartner.diagnostic_Console(self, 'Error', e)
            
    def Run_ResetECU_Func(self):
        try:
            TXID_PHYS  = int(self.Ter.TXID_lineEdit.text(),16)
            TXID_FUNC  = int('7DF',16)
            RXID       = int(self.Ter.RXID_lineEdit.text(),16) 
            REPORT_TYPE   = 'html'
            dt_string  = datetime.now().strftime("[%d-%m-%Y_%H.%M.%S]")
            Re_wb = Workbook()
            Re_ws = Re_wb.active
            #Active session DID:
            if self.Ter.AS_checkBox.isChecked():
                ACTIVE_SES = 'F186'
            elif (self.Ter.AS_checkBox.isChecked() == False) and (self.Ter.AS_lineEdit.text() in ''):
                SRPartner.diagnostic_Console(self, 'Error', 'AtiveSession DID can not be blank')
                return None
            elif (self.Ter.AS_checkBox.isChecked()) and (self.Ter.AS_lineEdit.text() not in ''):
                SRPartner.diagnostic_Console(self, 'Error', 'Only 1 AtiveSession DID acceptable')
                return None
            else:
                ACTIVE_SES = self.Ter.AS_lineEdit.text()
                
            #Check CAN FD or not:
            if self.Ter.CANFD_checkBox.isChecked(): ISCANFD = True
            else:                                   ISCANFD = False
            
            #Check address:
            if self.Ter.PhysDiag_checkBox.isChecked() and self.Ter.FuncDiag_checkBox.isChecked():
                SRPartner.diagnostic_Console(self, 'Error', 'Only Phys or Func diag address can be selected')
            elif (self.Ter.PhysDiag_checkBox.isChecked() == False) and (self.Ter.FuncDiag_checkBox.isChecked() == False):
                SRPartner.diagnostic_Console(self, 'Error', 'At least 1 ADDRESS should be checked')
            else:
                #Reset in Physical mode:
                if self.Ter.PhysDiag_checkBox.isChecked():
                    #Set address:
                    self.Ter.TXID_lineEdit.setText(str(hex(TXID_PHYS))[2:])
                    TXID = TXID_PHYS
                #Reset in Functional mode:
                elif self.Ter.FuncDiag_checkBox.isChecked():
                    #Set address:
                    self.Ter.TXID_lineEdit.setText(str(hex(TXID_FUNC))[2:])
                    TXID = TXID_FUNC
                else:
                    pass #For next release
                    
                #Check SPRMIB or not:
                if self.Ter.SPRMIB_checkBox.isChecked(): SPRMIB = True
                else:                                    SPRMIB = False
                
                #Check validation: Subfunction and Service
                if (self.Ter.DS_checkBox.isChecked() == False) and \
                    (self.Ter.PS_checkBox.isChecked() == False) and \
                    (self.Ter.ES_checkBox.isChecked() == False):
                    SRPartner.diagnostic_Console(self, 'Error', 'At least 1 SERVICE should be checked')
                else:
                    if (self.Ter.Sub01_checkBox.isChecked() == False) and \
                        (self.Ter.Sub02_checkBox.isChecked() == False) and \
                        (self.Ter.Sub03_checkBox.isChecked() == False):
                        SRPartner.diagnostic_Console(self, 'Error', 'At least 1 SUBFUNCTION should be checked')
                    else:
                        SER = {'DS' : self.is_DS_check, 'PS' : self.is_PS_check, 'ES' : self.is_ES_check}
                        SUB = {'Sub01': self.is_Sub01_check ,'Sub02' : self.is_Sub02_check ,'Sub03' : self.is_Sub03_check}
                        #Result:
                        RESULT = Func_INS.ResetECU_Func(TXID, RXID, ISCANFD, SPRMIB, SER, SUB, ACTIVE_SES)
            
            #Restore Physical Diag:
            self.Ter.TXID_lineEdit.setText(str(hex(TXID_PHYS))[2:])
            
            #Report:
            if REPORT_TYPE in 'Excel':
                #Excel:
                num = 2
                headings = ['No'] + ['TC_name'] + ['Diag'] + ['Request'] + ['Response']
                Re_ws.append(headings)
                
                for Re_key, Re_val in RESULT.items():
                    Re_ws[f'A{num}'] = Re_key
                    Re_ws[f'B{num}'] = Re_val['TC_name']
                    Re_ws[f'C{num}'] = Re_val['Diag']
                    length = len(Re_val['Request'])
                    for i in range(1,length+1):
                        Re_ws[f'D{num+i-1}'] = Re_val['Request'][i-1]
                        Re_ws[f'E{num+i-1}'] = Re_val['Response'][i-1]
                    num += length 
                        
                # Save the file 
                try :    Re_wb.save(f"{os.path.join(self.Ter.WFP_lineEdit.text(),'PIC', 'log', 'Test_log')}/Case_7_1_{dt_string}.xlsx")
                except : Re_wb.save(f"{self.Current_path}/temp/report/Case_7_1_{dt_string}.xlsx")
            else:
                #HTML:
                SRPartner.case_7_1_HTML(self,RESULT)
            SRPartner.diagnostic_Console(self, 'Success', 'Case 7.1 done!')
                        
        except Exception as e:
            SRPartner.diagnostic_Console(self, 'Error', e)
            
    def Run_Case_3_Func(self):
        # try: 
            dt_string = datetime.now().strftime("[%d-%m-%Y_%H.%M.%S]")
            path          = self.Ter.UDS_lineEdit.text()
            sheet         = self.Ter.UDSSheet_lineEdit.text()
            prj           = self.Ter.PRJ_comboBox.currentText()
            DID_cols      = 'A'
            DIDName_col   = 'B'
            TESTCASE      = '3'
            REPORT_TYPE   = 'html'
            TXID  = int(self.Ter.TXID_lineEdit.text(),16)
            RXID  = int(self.Ter.RXID_lineEdit.text(),16)
            ISCANFD = True
            #Check CAN FD or not:
            if self.Ter.CANFD_checkBox.isChecked(): ISCANFD = True
            else:                                   ISCANFD = False
            #Config:
            if path == '':
                SRPartner.diagnostic_Console(self, 'Error', 'Error UDS path.Please check')
            else:
                #Result:
                RESULT = Func_INS.Case_3_Func(path, sheet, DID_cols, DIDName_col, prj, TXID, RXID, ISCANFD)
                #report:
                if REPORT_TYPE in 'Excel':
                    #Excel:
                    df = pandas.DataFrame(RESULT).T
                    try:    df.to_excel(f"{os.path.join(self.Ter.WFP_lineEdit.text(),'PIC', 'log', 'Test_log')}/Case_3_{dt_string}.xlsx")
                    except: df.to_excel(f"{self.Current_path}/temp/report/Case_3_{dt_string}.xlsx")
                else:
                    #HTML:
                    SRPartner.HTML_Report(self,RESULT,TESTCASE)
                SRPartner.diagnostic_Console(self, 'Success', f'Case {TESTCASE} done!')    
        # except Exception as e:
        #     SRPartner.diagnostic_Console(self, 'Error', e)
        
    def Run_Oxflash_Browse_Func(self):
        path = str(self.LoadInit_dict['TabSRCheck']['SR_Filepath']['Oxflash_lineEdit']).replace('/','\\')
        if path == '':
            path = r'D:\.'
        OpenObject = QFileDialog.getExistingDirectory(parent= self, caption="Open sandbox folder", directory= path)
        if OpenObject == '':
            self.Ter.Oxflash_lineEdit.setText(self.Ter.Oxflash_lineEdit.text())
        else:
            self.Ter.Oxflash_lineEdit.setText(OpenObject)
        
    def Run_Nxflash_Browse_Func(self):
        path = str(self.LoadInit_dict['TabSRCheck']['SR_Filepath']['Nxflash_lineEdit']).replace('/','\\')
        if path == '':
            path = r'D:\.'
        OpenObject = QFileDialog.getExistingDirectory(parent= self, caption="Open sandbox folder", directory= path)
        if OpenObject == '':
            self.Ter.Nxflash_lineEdit.setText(self.Ter.Nxflash_lineEdit.text())
        else:
            self.Ter.Nxflash_lineEdit.setText(OpenObject)
        
    def Run_ESBBOX_Browse_Func(self):
        path = str(self.LoadInit_dict['TabSRCheck']['SR_Filepath']['ESBBOX_lineEdit']).replace('/','\\')
        if path == '':
            path = r'D:\.'
        OpenObject = QFileDialog.getExistingDirectory(parent= self, caption="Open ESBBOX folder", directory= path)
        if OpenObject == '':
            self.Ter.ESBBOX_lineEdit.setText(self.Ter.ESBBOX_lineEdit.text())
        else:
            self.Ter.ESBBOX_lineEdit.setText(OpenObject)
            
    def Run_STIL_Browse_Func(self):
        path = str(self.LoadInit_dict['TabSRCheck']['SR_Filepath']['STIL_lineEdit']).replace('/','\\')
        if path == '':
            path = r'D:\.'
        OpenObject = QFileDialog.getOpenFileName(parent= self,  caption="Open STIL folder", directory= path, filter = 'All (*.exe*)')
        if OpenObject[0] == '':
            self.Ter.STIL_lineEdit.setText(self.Ter.STIL_lineEdit.text())
        else:
            self.Ter.STIL_lineEdit.setText(OpenObject[0])
        SRPartner.UpdateJSON(self)
            
    def Run_FSI_SB_Browse_Func(self):
        path = str(self.LoadInit_dict['TabFSI']['FSI_Filepath']['FSI_SB_lineEdit']).replace('/','\\')
        if path == '':
            path = r'D:\.'
        OpenObject = QFileDialog.getExistingDirectory(parent= self, caption="Open sandbox folder", directory= path)
        if OpenObject == '':
            self.Ter.FSI_SB_lineEdit.setText(self.Ter.FSI_SB_lineEdit.text())
        else:
            self.Ter.FSI_SB_lineEdit.setText(OpenObject)
        SRPartner.UpdateJSON(self)
        
    def Run_UDS_Browse_Func(self):
        path = str(self.LoadInit_dict['TabSRCheck']['Control']['UDS_lineEdit']).replace('/','\\')
        if path == '':
            path = r'D:\.'
        OpenObject = QFileDialog.getOpenFileName(parent= self, caption="Open UDS file", directory= r'D:\.' , filter='All (*.xlsx*)' )
        if OpenObject[0] == '':
            self.Ter.UDS_lineEdit.setText(self.Ter.UDS_lineEdit.text())
        else:
            self.Ter.UDS_lineEdit.setText(OpenObject[0])
        SRPartner.UpdateJSON(self)
        
    def Run_FSI_Diatester_Browse_Func(self): 
        Func_INS.Tool_Browse_Func(self.Ter.FSI_SB_lineEdit, self.Ter.FSI_Diatester_lineEdit, 'All (*.cna*)')
        SRPartner.UpdateJSON(self)
        
    def Run_FSI_CANalyzer_Browse_Func(self):
        Func_INS.Tool_Browse_Func(self.Ter.FSI_SB_lineEdit, self.Ter.FSI_CANalyzer_lineEdit, 'All (*.cfg*)')
        SRPartner.UpdateJSON(self)
        
    def Run_FSI_CANape_Browse_Func(self):
        Func_INS.Tool_Browse_Func(self.Ter.FSI_SB_lineEdit, self.Ter.FSI_CANape_lineEdit, 'All (*.cna*)')
        SRPartner.UpdateJSON(self)
        
    def Run_FSI_FaultID_Browse_Func(self):
        Func_INS.Tool_Browse_Func(self.Ter.FSI_SB_lineEdit, self.Ter.FSI_FaultID_lineEdit, 'All (*.h*)')
        SRPartner.UpdateJSON(self)
        
    def Run_FSI_PubDBC_Browse_Func(self):
        Func_INS.Tool_Browse_Func(self.Ter.FSI_SB_lineEdit, self.Ter.FSI_PubDBC_lineEdit, 'All (*.dbc*)')
        SRPartner.UpdateJSON(self)
        
    def Run_FSI_PriDBC_Browse_Func(self):
        Func_INS.Tool_Browse_Func(self.Ter.FSI_SB_lineEdit, self.Ter.FSI_PriDBC_lineEdit, 'All (*.dbc*)')
        SRPartner.UpdateJSON(self)
        
    # -------------------- #ANCHOR: Generate buttons management --------------------:
    def Run_GenDocument_Func(self):
        try: 
            if self.Ter.WFP_lineEdit.text() == '' :
                SRPartner.Console(self,'Error', 'Invalid path') 
            elif self.Ter.Var_comboBox.currentText() == '':
                SRPartner.Console(self,'Error', 'Var cannot be empty') 
            elif self.Ter.Type_comboBox.currentText() == '':
                SRPartner.Console(self,'Error', 'Type cannot be empty') 
            elif self.Ter.PN_lineEdit.text() == '':
                SRPartner.Console(self,'Error', 'Part-Number cannot be empty') 
            elif self.Ter.BL_lineEdit.text() == '' or self.Ter.RCV_lineEdit.text() == '':
                SRPartner.Console(self,'Error', 'BL or RC/V cannot be empty') 
            elif (self.Ter.BL_lineEdit.text()[0] == '0' and len(self.Ter.BL_lineEdit.text()[0]) > 1) or \
                 (self.Ter.RCV_lineEdit.text()[0] == '0' and len(self.Ter.RCV_lineEdit.text()[0]) > 1):
                SRPartner.Console(self,'Error', 'Remove prefix \'0\' in BL or RC/V') 
            elif self.Ter.TTicket_lineEdit.text() == '':
                SRPartner.Console(self,'Error', 'JIRA cannot be empty') 
            else:
                ##########: Main code here ##########:
                #FR/CR:
                self.Pos = self.Ter.Pos_comboBox.currentText()

                #BL-RC-V:
                SRPartner.Config_SWVer(self)

                #Update global vars:
                self.Ter.SWVer_lineEdit.setText(self.SWVer)
                

                #Part-Number:   
                if str(self.Ter.PN_lineEdit.text())[0:4] == '0203' and len(str(self.Ter.PN_lineEdit.text())) == 10:
                    self.PN = self.Ter.PN_lineEdit.text()
                elif len(str(self.Ter.PN_lineEdit.text())) == 6:
                    self.PN = '0203{0}'.format(self.Ter.PN_lineEdit.text())
                else:
                    self.PN = "Invalid PN"

                #Gen Document:
                CHECKLIST,PROTOCOL = Func_INS.GenDocument_Func(path     = self.Ter.WFP_lineEdit.text(),
                                                               type     = self.Ter.Type_comboBox.currentText(),
                                                               prj      = self.Ter.Var_comboBox.currentText(),
                                                               SWver    = self.SWVer,
                                                               Pos      = self.Pos,
                                                               PN       = self.PN,
                                                               HWPCM    = self.Ter.HWPCM_lineEdit.text(),
                                                               SWPCM    = self.Ter.SWPCM_comboBox.currentText(),
                                                               TPM      = self.Ter.TPM_lineEdit.text(),
                                                               TC       = self.Ter.TC_comboBox.currentText(),
                                                               Tester   = self.Ter.Tester_comboBox.currentText(),
                                                               Reviewer = self.Ter.Reviewer_lineEdit.text(),
                                                               TM       = self.Ter.TM_lineEdit.text(),
                                                               SWpath   = self.Ter.SWP_lineEdit.text(),
                                                               ticket   = self.Ter.TTicket_lineEdit.text(),
                                                               )
                 
                #Update CL,Protocol name:
                self.CheckList_Name = str(CHECKLIST).split('/')[-1]
                self.Protocol_Name  = str(PROTOCOL).split('/')[-1]
                
                #Console:                               
                SRPartner.Console(self,'Satus', 'Generated Successfully')
            SRPartner.UpdateJSON(self)


        except FileExistsError:
            #Console:                               
            SRPartner.Console(self,'Error', 'Cannot create a file when that file already exists')
   
    def Run_SWCopy_Func(self):
        WFP = self.Ter.WFP_lineEdit.text()
        SWP = self.Ter.SWP_lineEdit.text()
        if SWP  in '':
            SRPartner.Console(self,'Error', 'SW path cannot be empty') 
        else:
            ##########: Main code here ##########:
            if 'New_SW' in os.listdir(WFP):
                #sub function:
                def SubFunc(path_x, SWpath_x):
                    path_x = str(path_x).replace("\\", "/")
                    SWpath_org = SWpath_x
                    SWpath_x = str(SWpath_x).replace("\\", "/")
                    #Copy SW to local:
                    for root, dirs, files in os.walk(SWpath_x):
                        for file in files:
                            if ('Series' in file) and ((file.endswith('.7z')) or (file.endswith('.zip'))):
                                # Source path
                                source = os.path.join(SWpath_org, file)
                                # Destination path
                                destination = os.path.join(path_x, 'New_SW')
                                # Copy the content of
                                # source to destination
                                shutil.copy(source, destination)
                                #Console:
                                SRPartner.Console(self,'Satus', 'Copy done')
                                #Extract SW
                                SRPartner.Console(self,'Satus', 'Extracting...')
                                filename = file.replace('.7z','').replace('.zip','')
                                destination_ZipPath = os.path.join(destination, file)
                                destination_ZipFolPath = os.path.join(destination, filename)
                                os.mkdir(destination_ZipFolPath)
                                if file.endswith('.7z'):
                                    Archive(destination_ZipPath).extractall(destination_ZipFolPath)
                                elif file.endswith('.zip'):
                                    with zipfile.ZipFile(destination_ZipPath, 'r') as zip_ref:
                                        zip_ref.extractall(destination_ZipFolPath)
                                #Console:
                                SRPartner.Console(self,'Satus', 'Extract done')
                                
                    if not any(os.scandir(destination)):
                        SRPartner.Console(self,'Error', 'No \'Series\' SW in the software path')

                    print('before done')
                    print('Done')
                t1 = threading.Thread(target=SubFunc, args=(WFP,SWP))
                t1.start()

            else:
                #Console:
                SRPartner.Console(self,'Error', 'The system cannot find the path specified \n[../New_SW]')

    def Run_GenReportBackup_Func(self):
        try:
            SWRCV = self.Ter.SWVer_lineEdit.text()  
            REPORTP = os.path.join(self.Ter.WFP_lineEdit.text(),'Final_Report')
            TICKET = self.Ter.TTicket_lineEdit.text()
            WFP = self.Ter.WFP_lineEdit.text()
            RET = []
            CL_NAME = str(self.LoadInit_dict['TabInfo']['Report']['CheckList_Name'])
            PROTOCOL_NAME = str(self.LoadInit_dict['TabInfo']['Report']['Protocol_Name'])
            CL_STATUS = Func_INS.CheckCL(os.path.join(self.Ter.WFP_lineEdit.text(),CL_NAME))
            if  CL_STATUS in 'failure' :
                SRPartner.Console(self,'Error', Config_INS.CL_check[CL_STATUS])
            else:  
                RET = Func_INS.GenReportBackup_Func(self.Ter.BKS_lineEdit.text(), SWRCV[5:], REPORTP, TICKET, WFP, PROTOCOL_NAME, CL_NAME, Config_INS.CL_check[CL_STATUS])
                #Console:
                SRPartner.Console(self,'Satus', 'Files Copied (Backup Server)')
                #Set init values:
                self.Ter.Report_BKS_lineEdit.setText(str(RET))
                SRPartner.UpdateJSON(self)
        except FileExistsError:
            #Console:
            SRPartner.Console(self,'Error', 'File already exist')
        except FileNotFoundError:
            #Console:
            SRPartner.Console(self,'Error', 'The system cannot find the path specified \n[../Final_Report]')
        except Exception as error:
            #Console:
            SRPartner.Console(self,'Error', error)

    def Run_GenReportSuzhou_Func(self):
        try:
            SWRCV = self.Ter.SWVer_lineEdit.text()
            REPORTP = os.path.join(self.Ter.WFP_lineEdit.text(),'Final_Report')
            TICKET = self.Ter.TTicket_lineEdit.text()
            WFP = self.Ter.WFP_lineEdit.text()
            RET = ['','']
            CL_NAME = str(self.LoadInit_dict['TabInfo']['Report']['CheckList_Name'])
            PROTOCOL_NAME = str(self.LoadInit_dict['TabInfo']['Report']['Protocol_Name'])
            CL_STATUS = Func_INS.CheckCL(os.path.join(self.Ter.WFP_lineEdit.text(),CL_NAME))
            if  CL_STATUS in 'failure' :
                SRPartner.Console(self,'Error', Config_INS.CL_check[CL_STATUS])
            else: 
                RET = Func_INS.GenReportSuzhou_Func(self.Ter.SS_lineEdit.text(), SWRCV[5:], REPORTP, TICKET, WFP, PROTOCOL_NAME, CL_NAME, Config_INS.CL_check[CL_STATUS])
                #Console:
                SRPartner.Console(self,'Satus', 'Files Copied (Suzhou Server)')
                #Set init values:
                self.Ter.Report_Protocol_lineEdit.setText(str(RET[1]).replace('\\','/'))
                self.Ter.Report_SS_lineEdit.setText(str(RET[0]))
                SRPartner.UpdateJSON(self)
        except FileExistsError:
            #Console:
            SRPartner.Console(self,'Error', 'File already exist')
        except FileNotFoundError:
            #Console:
            SRPartner.Console(self,'Error', 'The system cannot find the path specified [../Final_Report]')
        except Exception as error:
            #Console:
            SRPartner.Console(self,'Error', error)

    def Run_GenReportSRCheck_Func(self):
        try:
            SWRCV = self.Ter.SWVer_lineEdit.text()
            REPORTP = os.path.join(self.Ter.WFP_lineEdit.text(),'Final_Report')
            TICKET = self.Ter.TTicket_lineEdit.text()
            WFP = self.Ter.WFP_lineEdit.text()
            RET = []
            CL_NAME = str(self.LoadInit_dict['TabInfo']['Report']['CheckList_Name'])
            PROTOCOL_NAME = str(self.LoadInit_dict['TabInfo']['Report']['Protocol_Name'])
            CL_STATUS = Func_INS.CheckCL(os.path.join(self.Ter.WFP_lineEdit.text(),CL_NAME))
            if  CL_STATUS in 'failure' :
                SRPartner.Console(self,'Error', Config_INS.CL_check[CL_STATUS])
            else: 
                RET = Func_INS.GenReportSRCheck_Func(self.Ter.SRCR_lineEdit.text(), SWRCV, REPORTP, TICKET, WFP, PROTOCOL_NAME, CL_NAME, Config_INS.CL_check[CL_STATUS])
                #Console:
                SRPartner.Console(self,'Satus', 'Files Copied (SRCheck Server)')
                #Set init values:
                self.Ter.Report_SRCR_lineEdit.setText(str(RET))
                SRPartner.UpdateJSON(self)
        except FileExistsError:
            #Console:
            SRPartner.Console(self,'Error', 'File already exist')
        except FileNotFoundError:
            #Console:
            SRPartner.Console(self,'Error', 'The system cannot find the path specified [../Final_Report]')
        except Exception as error:
            #Console:
            SRPartner.Console(self,'Error', error)
        

    # -------------------- #ANCHOR: Open buttons management --------------------:
    #Info:
    def Run_SWBuildOpen_Func(self):
        Func_INS.SWBuildOpen_Func()

    def Run_BKSOpen_Func(self):
        Func_INS.BKSOpen_Func()

    def Run_SSOpen_Func(self):
        Func_INS.SSOpen_Func()

    def Run_SRCROpen_Func(self):
        Func_INS.SRCROpen_Func()
        
    def Run_LRFOpen_Func(self):
        Func_INS.RFOpen_Func(self.Ter.WFP_lineEdit.text())
        
    def Run_TLOpen_Func(self):
        Func_INS.TLOpen_Func(self.Ter.WFP_lineEdit.text())
        
    def Run_SWOpen_Func(self):
        try: 
            if self.Ter.SWP_lineEdit.text() == '':
                SRPartner.Console(self,'Error','SW path is empty')
            else:
                Func_INS.SWOpen_Func(self.Ter.SWP_lineEdit.text())
        except:
            SRPartner.Console(self,'Error','Invalid SW path')
        
    def Run_WFOpen_Func(self):
        try: 
            if self.Ter.WFP_lineEdit.text() == '':
                SRPartner.Console(self,'Error','WF path is empty')
            else:
                Func_INS.WFOpen_Func(self.Ter.WFP_lineEdit.text())
        except:
            SRPartner.Console(self,'Error','Invalid WF path')

    def Run_DBOpen_Func(self):
        path = self.Tooldatabase
        try: 
            if path == '':
                SRPartner.Console(self,'Error','Database path is empty')
            else:
                Func_INS.OpenFile(path.strip().replace(r'/', '\\'), 'txt')
        except:
            SRPartner.Console(self,'Error','Invalid Database path')

    def Run_SBOpen_Func(self):
        Func_INS.SBOpen_Func(self.Ter.SB_lineEdit.text())

    def Run_OxflashOpen_Func(self):
        Func_INS.xflashOpen_Func(self.Ter.Oxflash_lineEdit.text())
        
    def Run_NxflashOpen_Func(self):
        Func_INS.xflashOpen_Func(self.Ter.Nxflash_lineEdit.text())
        
    def Run_ESBBOXOpen_Func(self):
        Func_INS.ESBBOXOpen_Func(self.Ter.ESBBOX_lineEdit.text())
        
    def Run_STILOpen_Func(self):
        Func_INS.STILOpen_Func(self.Ter.STIL_lineEdit.text())
        
    def Run_TicketOpen_Func(self):
        Func_INS.TicketOpen_Func(self.Ter.TTicket_lineEdit.text())

    def Run_ReportBKSOpen_Func(self):
        path = self.Ter.Report_BKS_lineEdit.text()
        Func_INS.OpenFolder_Func(path)

    def Run_ReportSRCROpen_Func(self):
        path = self.Ter.Report_SRCR_lineEdit.text()
        Func_INS.OpenFolder_Func(path)

    def Run_ReportSSOpen_Func(self):
        path = self.Ter.Report_SS_lineEdit.text()
        Func_INS.OpenFolder_Func(path)

    def Run_ReportProtocolOpen_Func(self):
        path = self.Ter.Report_Protocol_lineEdit.text().replace(r'/', '\\').strip()
        Func_INS.OpenFile(path, 'excel')

    #FSI:
    def Run_FSI_SBOpen_Func(self):
        Func_INS.SBOpen_Func(self.Ter.FSI_SB_lineEdit.text())
        
    # SRCheck:
    def Run_SendDiagnostic_Func(self):
        try:
            TXID  = int(self.Ter.TXID_lineEdit.text(),16)
            RXID  = int(self.Ter.RXID_lineEdit.text(),16)
            CANFD = True 
            REQ   = self.Ter.Diagnostic_lineEdit.text()
            RES   = 'Invalid'
            if self.Ter.CANFD_checkBox.isChecked(): CANFD = True
            else:                                   CANFD = False
            SRPartner.diagnostic_Console(self,'Status','Request: {}'.format(REQ))
            RES = Func_INS.SendDiagnostic_Func(TXID, RXID, CANFD, REQ)
            #Console:
            if RES.__contains__('Error'):
                SRPartner.diagnostic_Console(self,'Error','Response: {}'.format(RES))
            else:
                SRPartner.diagnostic_Console(self,'Status','Response: {}'.format(RES))
        except Exception as error:
            SRPartner.diagnostic_Console(self,'Error',error)
    
    def Run_ESBBox_on_Func(self):
        Func_INS.ESBBox_on(self.Ter.ESBBOX_lineEdit.text(), self.Current_path)
        self.Ter.ESPON_Open_Button.setStyleSheet("background-color: lightgreen")
        self.Ter.ESPOFF_Open_Button.setStyleSheet("background-color: light gray")
        
    def Run_ESBBox_off_Func(self):
        Func_INS.ESBBox_off(self.Ter.ESBBOX_lineEdit.text(), self.Current_path)
        self.Ter.ESPON_Open_Button.setStyleSheet("background-color: light gray")
        self.Ter.ESPOFF_Open_Button.setStyleSheet("background-color: lightgreen")

    # -------------------- #ANCHOR: Combobox management --------------------:
    def Run_ChooseConvertMode_Func(self):
        if self.Ter.H2A_comboBox.currentText() == 'Hex to Ascii':
           self.Ter.H2A_HEX_label.setText("Hex:")
           self.Ter.H2A_ASCII_label.setText("Ascii:")
           self.Ter.H2A_groupBox.setTitle('Hex to Ascii')
           self.Ter.H2A_HEX_lineEdit.clear()
           self.Ter.H2A_ASCII_lineEdit.clear()
        elif self.Ter.H2A_comboBox.currentText() == 'Ascii to Hex':
           self.Ter.H2A_HEX_label.setText("Ascii:")
           self.Ter.H2A_ASCII_label.setText("Hex:")
           self.Ter.H2A_groupBox.setTitle('Ascii to Hex')
           self.Ter.H2A_HEX_lineEdit.clear()
           self.Ter.H2A_ASCII_lineEdit.clear()

    def Run_HexToAscii_Func(self):
        try:
            Converted_Value = Func_INS.HexToAscii_Func(mode=  self.Ter.H2A_comboBox.currentText(),
                                                    input= self.Ter.H2A_HEX_lineEdit.text())
            self.Ter.H2A_ASCII_lineEdit.setStyleSheet("color: black") #set text color
            self.Ter.H2A_ASCII_lineEdit.setText(Converted_Value)
        except ValueError: 
            self.Ter.H2A_ASCII_lineEdit.setStyleSheet("font-weight: bold; color: red") #set text color
            self.Ter.H2A_ASCII_lineEdit.setText('Invalid HEX value')
    
    def Run_ChooseVariant_Func(self):
        self.Ter.Var_comboBox.clear()
        Var_items = Func_INS.ChooseVariant_Func(self.Ter.PRJ_comboBox.currentText())
        self.Ter.Var_comboBox.addItems(Var_items)
        SRPartner.UpdateJSON(self)

    def Run_ChoosePosition_Func(self):
        SRPartner.UpdateJSON(self)

    def Run_ChooseRCV_Func(self):
        SRPartner.Config_SWVer(self)
        self.Ter.SWVer_lineEdit.setText(self.SWVer)
        SRPartner.UpdateJSON(self)

    def Run_ChooseVar_Func(self):
        SRPartner.UpdateJSON(self)

    def Run_ChooseType_Func(self):
        SRPartner.UpdateJSON(self)

    def Run_ChooseTester_Func(self):
        SRPartner.UpdateJSON(self)
        
    def Run_ChooseSWPCM_Func(self):
        SRPartner.UpdateJSON(self)
        
    def Run_ChooseTC_Func(self):
        SRPartner.UpdateJSON(self)

    # -------------------- #ANCHOR: Line Edit Changed --------------------:  
    def Run_WFP_lineEditChanged_Func(self,):
        SRPartner.UpdateJSON(self)

    def Run_SWP_lineEditChanged_Func(self):
        SRPartner.UpdateJSON(self)

    def Run_PN_lineEditChanged_Func(self):
        SRPartner.UpdateJSON(self)

    def Run_BL_lineEditChanged_Func(self):
        SRPartner.Config_SWVer(self)
        self.Ter.SWVer_lineEdit.setText(self.SWVer)
        SRPartner.UpdateJSON(self)

    def Run_RCV_lineEditChanged_Func(self):
        SRPartner.Config_SWVer(self)
        self.Ter.SWVer_lineEdit.setText(self.SWVer)
        SRPartner.UpdateJSON(self)

    def Run_HWPCM_lineEditChanged_Func(self):
        SRPartner.UpdateJSON(self)
        
    def Run_TPM_lineEditChanged_Func(self):
        SRPartner.UpdateJSON(self)
        
    def Run_BKS_lineEditChanged_Func(self):
        SRPartner.UpdateJSON(self)

    def Run_SS_lineEditChanged_Func(self):
        SRPartner.UpdateJSON(self)

    def Run_SRCR_lineEditChanged_Func(self):
        SRPartner.UpdateJSON(self)

    def Run_TTicket_lineEditChanged_Func(self):
        SRPartner.UpdateJSON(self)

    def Run_Reviewer_lineEditChanged_Func(self):
        SRPartner.UpdateJSON(self)

    def Run_TM_lineEditChanged_Func(self):
        SRPartner.UpdateJSON(self)

    def Run_Report_BKS_lineEditChanged_Func(self):
        SRPartner.UpdateJSON(self)

    def Run_Report_SRCR_lineEditChanged_Func(self):
        SRPartner.UpdateJSON(self)

    def Run_Report_SS_lineEditChanged_Func(self):
        SRPartner.UpdateJSON(self)

    def Run_Report_Protocol_lineEditChanged_Func(self):
        SRPartner.UpdateJSON(self)

    def Run_SWVer_lineEditChanged_Func(self):
        SRPartner.UpdateJSON(self)


    def Run_SB_lineEditChanged_Func(self):
        try:
            Var = self.Ter.Var_comboBox.currentText()
            Var = str(Var).split('_')[1]
            Func_INS.Fulfill_ToolSRCheck_Func(self.Ter.SB_lineEdit.text(),
                                            self.Ter.Diatester_lineEdit,
                                            self.Ter.CANalyzer_lineEdit,
                                            self.Ter.CANape_lineEdit,
                                            self.Ter.FaultID_lineEdit,
                                            self.Ter.PubDBC_lineEdit,
                                            self.Ter.PriDBC_lineEdit,
                                            self.Ter.Var_comboBox.currentText())
            SRPartner.UpdateJSON(self)
        except:
            #Console:
            SRPartner.Console(self,'Error', 'Invalid Sandbox Path')
            
    def Run_FSI_SB_lineEditChanged_Func(self):
        try:
            Var = self.Ter.Var_comboBox.currentText()
            Var = str(Var).split('_')[1]
            Func_INS.Fulfill_ToolSRCheck_Func(self.Ter.FSI_SB_lineEdit.text(),
                                            self.Ter.FSI_Diatester_lineEdit,
                                            self.Ter.FSI_CANalyzer_lineEdit,
                                            self.Ter.FSI_CANape_lineEdit,
                                            self.Ter.FSI_FaultID_lineEdit,
                                            self.Ter.FSI_PubDBC_lineEdit,
                                            self.Ter.FSI_PriDBC_lineEdit,
                                            self.Ter.Var_comboBox.currentText())
            SRPartner.UpdateJSON(self)
        except:
            #Console:
            SRPartner.Console(self,'Error', 'Invalid Sandbox Path')
    
    def Run_Diatester_lineEditChanged_Func(self):
        SRPartner.UpdateJSON(self)

    def Run_CANalyzer_lineEditChanged_Func(self):
        SRPartner.UpdateJSON(self)

    def Run_CANape_lineEditChanged_Func(self):
        SRPartner.UpdateJSON(self)

    def Run_FaultID_lineEditChanged_Func(self):
        SRPartner.UpdateJSON(self)

    def Run_PubDBC_lineEditChanged_Func(self):
        SRPartner.UpdateJSON(self)

    def Run_PriDBC_lineEditChanged_Func(self):
        SRPartner.UpdateJSON(self)

    def Run_Oxflash_lineEditChanged_Func(self):
        SRPartner.UpdateJSON(self)

    def Run_Nxflash_lineEditChanged_Func(self):
        SRPartner.UpdateJSON(self)
        
    def Run_ESBBOX_lineEditChanged_Func(self):
        SRPartner.UpdateJSON(self)
        
    def Run_STIL_lineEditChanged_Func(self):
        SRPartner.UpdateJSON(self)
        
    # Control:
    def Run_TXID_lineEditChanged_Func(self):
        SRPartner.UpdateJSON(self)
        
    def Run_RXID_lineEditChanged_Func(self):
        SRPartner.UpdateJSON(self)
        
    def Run_AS_lineEditChanged_Func(self):
        SRPartner.UpdateJSON(self)
        
    def Run_UDS_lineEditChanged_Func(self):
        SRPartner.UpdateJSON(self)
        
    def Run_UDSSheet_lineEditChanged_Func(self):
        SRPartner.UpdateJSON(self)
        
    def Run_Diagnostic_lineEditChanged_Func(self):
        SRPartner.UpdateJSON(self)

    # -------------------- #ANCHOR: Open Vecter, FaultID, ClaraTE --------------------:
    #SR-Check:
    def Run_DiatesterOpen_Func(self):
        self.Ter.Diatester_Open_Button.setEnabled(False)
        QTimer.singleShot(2000, lambda: self.Ter.Diatester_Open_Button.setEnabled(True))
        if self.Ter.Diatester_lineEdit.text().strip() == '':
            tDir=''
            if self.Ter.SB_lineEdit.text().strip() != '':
                tDir = self.Ter.SB_lineEdit.text().strip()
            else:
                tDir = 'D:\\'
            file_filter = 'All (*.cna*)'
            file_object = QFileDialog.getOpenFileName( parent=self, caption='Select a data file', directory=tDir, filter=file_filter)
            self.Ter.Diatester_lineEdit.setText(file_object[0])
            SRPartner.UpdateJSON(self)
        else:
            Func_INS.OpenFile(self.Ter.Diatester_lineEdit.text().strip().replace(r'/', '\\'), 'cna')
        
    def Run_CANapeOpen_Func(self):
        self.Ter.CANape_Open_Button.setEnabled(False)
        QTimer.singleShot(2000, lambda: self.Ter.CANape_Open_Button.setEnabled(True))
        if self.Ter.CANape_lineEdit.text().strip() == '':
            tDir=''
            if self.Ter.SB_lineEdit.text().strip() != '':
                tDir = self.Ter.SB_lineEdit.text().strip()
            else:
                tDir = 'D:\\'
            file_filter = 'All (*.cna*)'
            file_object = QFileDialog.getOpenFileName( parent=self, caption='Select a data file', directory=tDir, filter=file_filter)
            self.Ter.CANape_lineEdit.setText(file_object[0])
            SRPartner.UpdateJSON(self)
        else:
            Func_INS.OpenFile(self.Ter.CANape_lineEdit.text().strip().replace(r'/', '\\'), 'cna')

    def Run_CANalyzerOpen_Func(self):
        self.Ter.CANalyzer_Open_Button.setEnabled(False)
        QTimer.singleShot(2000, lambda: self.Ter.CANalyzer_Open_Button.setEnabled(True))
        if self.Ter.CANalyzer_lineEdit.text().strip() == '':
            tDir=''
            if self.Ter.SB_lineEdit.text().strip() != '':
                tDir = self.Ter.SB_lineEdit.text().strip()
            else:
                tDir = 'D:\\'
            file_filter = 'All (*.cfg*)'
            file_object = QFileDialog.getOpenFileName( parent=self, caption='Select a data file', directory=tDir, filter=file_filter)
            self.Ter.CANalyzer_lineEdit.setText(file_object[0])
            SRPartner.UpdateJSON(self)
        else:
            Func_INS.OpenFile(self.Ter.CANalyzer_lineEdit.text().strip().replace(r'/', '\\'), 'cfg')
            
    def Run_FaultIDOpen_Func(self):
        self.Ter.FaultID_Open_Button.setEnabled(False)
        QTimer.singleShot(2000, lambda: self.Ter.FaultID_Open_Button.setEnabled(True))
        if self.Ter.FaultID_lineEdit.text().strip() == '':
            tDir=''
            if self.Ter.SB_lineEdit.text().strip() != '':
                tDir = self.Ter.SB_lineEdit.text().strip()
            else:
                tDir = 'D:\\'
            file_filter = 'All (*.h*)'
            file_object = QFileDialog.getOpenFileName( parent=self, caption='Select a data file', directory=tDir, filter=file_filter)
            self.Ter.FaultID_lineEdit.setText(file_object[0])
            SRPartner.UpdateJSON(self)
        else:
            Func_INS.OpenFile(self.Ter.FaultID_lineEdit.text().strip().replace(r'/', '\\'), 'txt')
            
    def Run_PubDBCOpen_Func(self):
        self.Ter.PubDBC_Open_Button.setEnabled(False)
        QTimer.singleShot(2000, lambda: self.Ter.PubDBC_Open_Button.setEnabled(True))
        if self.Ter.PubDBC_lineEdit.text().strip() == '':
            tDir=''
            if self.Ter.SB_lineEdit.text().strip() != '':
                tDir = self.Ter.SB_lineEdit.text().strip()
            else:
                tDir = 'D:\\'
            file_filter = 'All (*.dbc*)'
            file_object = QFileDialog.getOpenFileName( parent=self, caption='Select a data file', directory=tDir, filter=file_filter)
            self.Ter.PubDBC_lineEdit.setText(file_object[0])
            SRPartner.UpdateJSON(self)
        else:
            Func_INS.OpenFile(self.Ter.PubDBC_lineEdit.text().strip().replace(r'/', '\\'), 'dbc')
            
    def Run_PriDBCOpen_Func(self):
        self.Ter.PriDBC_Open_Button.setEnabled(False)
        QTimer.singleShot(2000, lambda: self.Ter.PriDBC_Open_Button.setEnabled(True))
        if self.Ter.PriDBC_lineEdit.text().strip() == '':
            tDir=''
            if self.Ter.SB_lineEdit.text().strip() != '':
                tDir = self.Ter.SB_lineEdit.text().strip()
            else:
                tDir = 'D:\\'
            file_filter = 'All (*.dbc*)'
            file_object = QFileDialog.getOpenFileName( parent=self, caption='Select a data file', directory=tDir, filter=file_filter)
            self.Ter.PriDBC_lineEdit.setText(file_object[0])
            SRPartner.UpdateJSON(self)
        else:
            Func_INS.OpenFile(self.Ter.PriDBC_lineEdit.text().strip().replace(r'/', '\\'), 'dbc')
            
    #FSI:
    def Run_FSI_DiatesterOpen_Func(self):
        self.Ter.FSI_Diatester_Open_Button.setEnabled(False)
        QTimer.singleShot(2000, lambda: self.Ter.FSI_Diatester_Open_Button.setEnabled(True))
        if self.Ter.FSI_Diatester_lineEdit.text().strip() == '':
            tDir=''
            if self.Ter.FSI_SB_lineEdit.text().strip() != '':
                tDir = self.Ter.FSI_SB_lineEdit.text().strip()
            else:
                tDir = 'D:\\'
            file_filter = 'All (*.cna*)'
            file_object = QFileDialog.getOpenFileName( parent=self, caption='Select a data file', directory=tDir, filter=file_filter)
            self.Ter.FSI_Diatester_lineEdit.setText(file_object[0])
            SRPartner.UpdateJSON(self)
        else:
            Func_INS.OpenFile(self.Ter.FSI_Diatester_lineEdit.text().strip().replace(r'/', '\\'), 'cna')
        
    def Run_FSI_CANapeOpen_Func(self):
        self.Ter.FSI_CANape_Open_Button.setEnabled(False)
        QTimer.singleShot(2000, lambda: self.Ter.FSI_CANape_Open_Button.setEnabled(True))
        if self.Ter.FSI_CANape_lineEdit.text().strip() == '':
            tDir=''
            if self.Ter.FSI_SB_lineEdit.text().strip() != '':
                tDir = self.Ter.FSI_SB_lineEdit.text().strip()
            else:
                tDir = 'D:\\'
            file_filter = 'All (*.cna*)'
            file_object = QFileDialog.getOpenFileName( parent=self, caption='Select a data file', directory=tDir, filter=file_filter)
            self.Ter.FSI_CANape_lineEdit.setText(file_object[0])
            SRPartner.UpdateJSON(self)
        else:
            Func_INS.OpenFile(self.Ter.FSI_CANape_lineEdit.text().strip().replace(r'/', '\\'), 'cna')

    def Run_FSI_CANalyzerOpen_Func(self):
        self.Ter.FSI_CANalyzer_Open_Button.setEnabled(False)
        QTimer.singleShot(2000, lambda: self.Ter.FSI_CANalyzer_Open_Button.setEnabled(True))
        if self.Ter.FSI_CANalyzer_lineEdit.text().strip() == '':
            tDir=''
            if self.Ter.FSI_SB_lineEdit.text().strip() != '': 
                tDir = self.Ter.FSI_SB_lineEdit.text().strip()
            else:
                tDir = 'D:\\'
            file_filter = 'All (*.cfg*)'
            file_object = QFileDialog.getOpenFileName( parent=self, caption='Select a data file', directory=tDir, filter=file_filter)
            self.Ter.FSI_CANalyzer_lineEdit.setText(file_object[0])
            SRPartner.UpdateJSON(self)
        else:
            Func_INS.OpenFile(self.Ter.FSI_CANalyzer_lineEdit.text().strip().replace(r'/', '\\'), 'cfg')
            
    def Run_FSI_FaultIDOpen_Func(self):
        self.Ter.FSI_FaultID_Open_Button.setEnabled(False)
        QTimer.singleShot(2000, lambda: self.Ter.FSI_FaultID_Open_Button.setEnabled(True))
        if self.Ter.FSI_FaultID_lineEdit.text().strip() == '':
            tDir=''
            if self.Ter.FSI_SB_lineEdit.text().strip() != '':
                tDir = self.Ter.FSI_SB_lineEdit.text().strip()
            else:
                tDir = 'D:\\'
            file_filter = 'All (*.h*)'
            file_object = QFileDialog.getOpenFileName( parent=self, caption='Select a data file', directory=tDir, filter=file_filter)
            self.Ter.FSI_FaultID_lineEdit.setText(file_object[0])
            SRPartner.UpdateJSON(self)
        else:
            Func_INS.OpenFile(self.Ter.FSI_FaultID_lineEdit.text().strip().replace(r'/', '\\'), 'txt')
            
    def Run_FSI_PubDBCOpen_Func(self):
        self.Ter.FSI_PubDBC_Open_Button.setEnabled(False)
        QTimer.singleShot(2000, lambda: self.Ter.FSI_PubDBC_Open_Button.setEnabled(True))
        if self.Ter.FSI_PubDBC_lineEdit.text().strip() == '':
            tDir=''
            if self.Ter.FSI_SB_lineEdit.text().strip() != '':
                tDir = self.Ter.FSI_SB_lineEdit.text().strip()
            else:
                tDir = 'D:\\'
            file_filter = 'All (*.dbc*)'
            file_object = QFileDialog.getOpenFileName( parent=self, caption='Select a data file', directory=tDir, filter=file_filter)
            self.Ter.FSI_PubDBC_lineEdit.setText(file_object[0])
            SRPartner.UpdateJSON(self)
        else:
            Func_INS.OpenFile(self.Ter.FSI_PubDBC_lineEdit.text().strip().replace(r'/', '\\'), 'dbc')
            
    def Run_FSI_PriDBCOpen_Func(self):
        self.Ter.FSI_PriDBC_Open_Button.setEnabled(False)
        QTimer.singleShot(2000, lambda: self.Ter.FSI_PriDBC_Open_Button.setEnabled(True))
        if self.Ter.FSI_PriDBC_lineEdit.text().strip() == '':
            tDir=''
            if self.Ter.FSI_SB_lineEdit.text().strip() != '':
                tDir = self.Ter.FSI_SB_lineEdit.text().strip()
            else:
                tDir = 'D:\\'
            file_filter = 'All (*.dbc*)'
            file_object = QFileDialog.getOpenFileName( parent=self, caption='Select a data file', directory=tDir, filter=file_filter)
            self.Ter.FSI_PriDBC_lineEdit.setText(file_object[0])
            SRPartner.UpdateJSON(self)
        else:
            Func_INS.OpenFile(self.Ter.FSI_PriDBC_lineEdit.text().strip().replace(r'/', '\\'), 'dbc')
            
    def Run_ClaraTEOpen_Func(self):
        self.Ter.ClaraTE_Open_Button.setEnabled(False)
        QTimer.singleShot(2000, lambda: self.Ter.ClaraTE_Open_Button.setEnabled(True))
        curdir = os.getcwd()
        if self.Ter.ClaraTE_lineEdit.text() == '':
            file_filter = 'All (*.exe*)'
            file_object = QFileDialog.getOpenFileName( parent=self, caption='Select ClaraTE.exe file', directory = 'C:\\' ,filter=file_filter)
            self.Ter.ClaraTE_lineEdit.setText(file_object[0])
            self.UpdateJSON()
        else:
            if os.path.isfile(self.Ter.ClaraTE_lineEdit.text().strip()):
                path, filename = os.path.split(self.Ter.ClaraTE_lineEdit.text().strip().replace(r'/', '\\'))
                os.chdir(path)
                try:
                    subprocess.Popen(filename)
                    ### implement close and re-open, while opening user click open again
                except Exception as e:
                    QMessageBox.critical(self, 'Warning', str(e))
                os.chdir(curdir)
            else:
                QMessageBox.critical(self, 'Warning', 'Can not found ClaraTE.exe')        
            
    def Run_OpenFlashCANape_Func(self):
        path = r'D:\GEN5_FSI\00_FLASH_Canape\RadarFC.cna'.replace(r'/', '\\')
        Func_INS.OpenFile(path, 'cna')
        
    # -------------------- #ANCHOR: Clear text --------------------:
    def Run_WFPTextClear_Func(self):
        self.Ter.WFP_lineEdit.clear()

    def Run_SWPTextClear_Func(self):
        self.Ter.SWP_lineEdit.clear()

    def Run_BKSTextClear_Func(self):
        self.Ter.BKS_lineEdit.clear()

    def Run_SSTextClear_Func(self):
        self.Ter.SS_lineEdit.clear()

    def Run_PNTextClear_Func(self):
        self.Ter.PN_lineEdit.clear()

    def Run_SRCRTextClear_Func(self):
        self.Ter.SRCR_lineEdit.clear()
        
    # -------------------- #ANCHOR: Console Management --------------------:
    def Console(self, type, context):
        now = datetime.now()
        dt_string = now.strftime("[%d-%m-%Y %H:%M:%S]")
        if type == 'Error':
            self.Ter.Console_textBrowser.append(f"<p style='color:red'>{'{0}: Error - {1}'.format(dt_string, context)}</p>")
        elif type == 'Satus':
            self.Ter.Console_textBrowser.append(f"<p style='color:black'>{'{0}: {1}'.format(dt_string, context)}</p>")
        else:
            pass
        
    def diagnostic_Console(self, type, context):
        now = datetime.now()
        dt_string = now.strftime("[%d-%m-%Y %H:%M:%S]")
        if type == 'Error':
            self.Ter.Diagnostic_Console.append(f"<p style='color:red'>{'{0}: Error - {1}'.format(dt_string, context)}</p>")
        elif type == 'Status':
            self.Ter.Diagnostic_Console.append(f"<p style='color:black'>{'{0}: {1}'.format(dt_string, context)}</p>")
        elif type == 'Success':
            self.Ter.Diagnostic_Console.append(f"<p style='color:green'>{'{0}: {1}'.format(dt_string, context)}</p>")
        else:
            pass

    # -------------------- #ANCHOR: JSON --------------------:
    # ANCHOR - Load Init Value
    def LoadJSON(self):
        # ANCHOR - LoadJSON - TabInfo ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
        self.Ter.Info_tabWidget.setCurrentIndex(int(self.LoadInit_dict['Tab']['Info_tabWidget']['CurrentIndex']))
        # Testware:
        self.Ter.WFP_lineEdit.setText(str(self.LoadInit_dict['TabInfo']['Testware']['WFP_lineEdit']).replace('/','\\'))
        self.Ter.SWP_lineEdit.setText(str(self.LoadInit_dict['TabInfo']['Testware']['SWP_lineEdit']).replace('/','\\'))
        self.Ter.PRJ_comboBox.setCurrentText(self.LoadInit_dict['TabInfo']['Testware']['PRJ_comboBox'])
        self.Ter.Type_comboBox.setCurrentText(self.LoadInit_dict['TabInfo']['Testware']['Type_comboBox'])
        self.Ter.Pos_comboBox.setCurrentText(self.LoadInit_dict['TabInfo']['Testware']['Position'])
        self.Ter.RCV_comboBox.setCurrentText(self.LoadInit_dict['TabInfo']['Testware']['RCV_comboBox'])
        self.Ter.BL_lineEdit.setText(self.LoadInit_dict['TabInfo']['Testware']['BL_lineEdit'])
        self.Ter.RCV_lineEdit.setText(self.LoadInit_dict['TabInfo']['Testware']['RCV_lineEdit'])
        self.Ter.SWVer_lineEdit.setText(self.LoadInit_dict['TabInfo']['Testware']['SWVer_lineEdit'])
        self.Ter.PN_lineEdit.setText(self.LoadInit_dict['TabInfo']['Testware']['PN_lineEdit'])
        self.Ter.HWPCM_lineEdit.setText(self.LoadInit_dict['TabInfo']['Testware']['HWPCM_lineEdit'])
        self.Ter.Tester_comboBox.setCurrentText(self.LoadInit_dict['TabInfo']['Testware']['Tester_comboBox'])
        self.Ter.SWPCM_comboBox.setCurrentText(self.LoadInit_dict['TabInfo']['Testware']['SWPCM_comboBox'])
        self.Ter.TC_comboBox.setCurrentText(self.LoadInit_dict['TabInfo']['Testware']['TC_comboBox'])
        self.Ter.TPM_lineEdit.setText(self.LoadInit_dict['TabInfo']['Testware']['TPM_lineEdit'])
        # Report:
        self.Ter.Report_SS_lineEdit.setText(str(self.LoadInit_dict['TabInfo']['Report']['Report_SS_lineEdit']).replace('/','\\'))
        self.Ter.Report_Protocol_lineEdit.setText(str(self.LoadInit_dict['TabInfo']['Report']['Report_Protocol_lineEdit']).replace('/','\\'))
        self.Ter.Report_BKS_lineEdit.setText(str(self.LoadInit_dict['TabInfo']['Report']['Report_BKS_lineEdit']).replace('/','\\'))
        self.Ter.Report_SRCR_lineEdit.setText(str(self.LoadInit_dict['TabInfo']['Report']['Report_SRCR_lineEdit']).replace('/','\\'))
        self.Ter.BKS_lineEdit.setText(str(self.LoadInit_dict['TabInfo']['Report']['BKS_lineEdit']).replace('/','\\'))
        self.Ter.SS_lineEdit.setText(str(self.LoadInit_dict['TabInfo']['Report']['SS_lineEdit']).replace('/','\\'))
        self.Ter.SRCR_lineEdit.setText(str(self.LoadInit_dict['TabInfo']['Report']['SRCR_lineEdit']).replace('/','\\'))
        self.Ter.TTicket_lineEdit.setText(self.LoadInit_dict['TabInfo']['Report']['TTicket_lineEdit'])
        self.Ter.Reviewer_lineEdit.setText(self.LoadInit_dict['TabInfo']['Report']['Reviewer_lineEdit'])
        self.Ter.TM_lineEdit.setText(self.LoadInit_dict['TabInfo']['Report']['TM_lineEdit'])
        # ANCHOR - LoadJSON - TabSRCheck ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
        # SR_Filepath:
        self.Ter.SB_lineEdit.setText(str(self.LoadInit_dict['TabSRCheck']['SR_Filepath']['SB_lineEdit']).replace('/','\\'))
        self.Ter.Diatester_lineEdit.setText(str(self.LoadInit_dict['TabSRCheck']['SR_Filepath']['Diatester_lineEdit']).replace('/','\\'))
        self.Ter.CANalyzer_lineEdit.setText(str(self.LoadInit_dict['TabSRCheck']['SR_Filepath']['CANalyzer_lineEdit']).replace('/','\\'))
        self.Ter.CANape_lineEdit.setText(str(self.LoadInit_dict['TabSRCheck']['SR_Filepath']['CANape_lineEdit']).replace('/','\\'))
        self.Ter.FaultID_lineEdit.setText(str(self.LoadInit_dict['TabSRCheck']['SR_Filepath']['FaultID_lineEdit']).replace('/','\\'))
        self.Ter.PubDBC_lineEdit.setText(str(self.LoadInit_dict['TabSRCheck']['SR_Filepath']['PubDBC_lineEdit']).replace('/','\\'))
        self.Ter.PriDBC_lineEdit.setText(str(self.LoadInit_dict['TabSRCheck']['SR_Filepath']['PriDBC_lineEdit']).replace('/','\\'))
        self.Ter.Oxflash_lineEdit.setText(str(self.LoadInit_dict['TabSRCheck']['SR_Filepath']['Oxflash_lineEdit']).replace('/','\\'))
        self.Ter.Nxflash_lineEdit.setText(str(self.LoadInit_dict['TabSRCheck']['SR_Filepath']['Nxflash_lineEdit']).replace('/','\\'))
        self.Ter.ESBBOX_lineEdit.setText(str(self.LoadInit_dict['TabSRCheck']['SR_Filepath']['ESBBOX_lineEdit']).replace('/','\\'))
        self.Ter.STIL_lineEdit.setText(str(self.LoadInit_dict['TabSRCheck']['SR_Filepath']['STIL_lineEdit']).replace('/','\\'))
        # Control
        self.Ter.TXID_lineEdit.setText(self.LoadInit_dict['TabSRCheck']['Control']['TXID_lineEdit'])
        self.Ter.RXID_lineEdit.setText(self.LoadInit_dict['TabSRCheck']['Control']['RXID_lineEdit'])
        self.Ter.AS_lineEdit.setText(self.LoadInit_dict['TabSRCheck']['Control']['AS_lineEdit'])
        self.Ter.Diagnostic_lineEdit.setText(self.LoadInit_dict['TabSRCheck']['Control']['Diagnostic_lineEdit'])
        self.Ter.UDS_lineEdit.setText(str(self.LoadInit_dict['TabSRCheck']['Control']['UDS_lineEdit']).replace('/','\\'))
        self.Ter.UDSSheet_lineEdit.setText(self.LoadInit_dict['TabSRCheck']['Control']['UDSSheet_lineEdit'])
        self.Ter.AS_checkBox.setChecked(bool(self.LoadInit_dict['TabSRCheck']['Control']['AS_checkBox']))
        self.Ter.DS_checkBox.setChecked(bool(self.LoadInit_dict['TabSRCheck']['Control']['DS_checkBox']))
        self.Ter.PS_checkBox.setChecked(bool(self.LoadInit_dict['TabSRCheck']['Control']['PS_checkBox']))
        self.Ter.ES_checkBox.setChecked(bool(self.LoadInit_dict['TabSRCheck']['Control']['ES_checkBox']))
        self.Ter.PhysDiag_checkBox.setChecked(bool(self.LoadInit_dict['TabSRCheck']['Control']['PhysDiag_checkBox']))
        self.Ter.FuncDiag_checkBox.setChecked(bool(self.LoadInit_dict['TabSRCheck']['Control']['FuncDiag_checkBox']))
        self.Ter.Sub01_checkBox.setChecked(bool(self.LoadInit_dict['TabSRCheck']['Control']['Sub01_checkBox']))
        self.Ter.Sub02_checkBox.setChecked(bool(self.LoadInit_dict['TabSRCheck']['Control']['Sub02_checkBox']))
        self.Ter.Sub03_checkBox.setChecked(bool(self.LoadInit_dict['TabSRCheck']['Control']['Sub03_checkBox']))
        self.Ter.SPRMIB_checkBox.setChecked(bool(self.LoadInit_dict['TabSRCheck']['Control']['SPRMIB_checkBox']))
        # ANCHOR - LoadJSON - TabSRCheck ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
        # SR_FSI
        self.Ter.FSI_SB_lineEdit.setText(str(self.LoadInit_dict['TabFSI']['FSI_Filepath']['FSI_SB_lineEdit']).replace('/','\\'))
        self.Ter.FSI_Diatester_lineEdit.setText(str(self.LoadInit_dict['TabFSI']['FSI_Filepath']['FSI_Diatester_lineEdit']).replace('/','\\'))
        self.Ter.FSI_CANalyzer_lineEdit.setText(str(self.LoadInit_dict['TabFSI']['FSI_Filepath']['FSI_CANalyzer_lineEdit']).replace('/','\\'))
        self.Ter.FSI_CANape_lineEdit.setText(str(self.LoadInit_dict['TabFSI']['FSI_Filepath']['FSI_CANape_lineEdit']).replace('/','\\'))
        self.Ter.FSI_FaultID_lineEdit.setText(str(self.LoadInit_dict['TabFSI']['FSI_Filepath']['FSI_FaultID_lineEdit']).replace('/','\\'))
        self.Ter.FSI_PubDBC_lineEdit.setText(str(self.LoadInit_dict['TabFSI']['FSI_Filepath']['FSI_PubDBC_lineEdit']).replace('/','\\'))
        self.Ter.FSI_PriDBC_lineEdit.setText(str(self.LoadInit_dict['TabFSI']['FSI_Filepath']['FSI_PriDBC_lineEdit']).replace('/','\\'))
        self.Ter.ClaraTE_lineEdit.setText(str(self.LoadInit_dict['TabFSI']['FSI_Filepath']['ClaraTE_lineEdit']).replace('/','\\'))
        
        #Init variant:
        self.Ter.Var_comboBox.clear()
        Var_items = Func_INS.ChooseVariant_Func(self.Ter.PRJ_comboBox.currentText())
        self.Ter.Var_comboBox.addItems(Var_items)
        self.Ter.Var_comboBox.setCurrentText(self.LoadInit_dict['TabInfo']['Testware']['Var_comboBox'])
        
        #Checkbox:
        self.is_ActSes_F186    = bool(self.LoadInit_dict['TabSRCheck']['Control']['AS_checkBox'])
        self.is_DS_check       = bool(self.LoadInit_dict['TabSRCheck']['Control']['DS_checkBox'])
        self.is_PS_check       = bool(self.LoadInit_dict['TabSRCheck']['Control']['PS_checkBox'])
        self.is_ES_check       = bool(self.LoadInit_dict['TabSRCheck']['Control']['ES_checkBox'])
        self.is_PhysDiag_check = bool(self.LoadInit_dict['TabSRCheck']['Control']['PhysDiag_checkBox'])
        self.is_FuncDiag_check = bool(self.LoadInit_dict['TabSRCheck']['Control']['FuncDiag_checkBox'])
        self.is_Sub01_check    = bool(self.LoadInit_dict['TabSRCheck']['Control']['Sub01_checkBox'])
        self.is_Sub02_check    = bool(self.LoadInit_dict['TabSRCheck']['Control']['Sub02_checkBox'])
        self.is_Sub03_check    = bool(self.LoadInit_dict['TabSRCheck']['Control']['Sub03_checkBox'])
        self.is_SPRMIB_check   = bool(self.LoadInit_dict['TabSRCheck']['Control']['SPRMIB_checkBox'])
        self.CheckList_Name    = str(self.LoadInit_dict['TabInfo']['Report']['CheckList_Name'])
        self.Protocol_Name     = str(self.LoadInit_dict['TabInfo']['Report']['Protocol_Name'])
        
        #Enable:
        if self.is_ActSes_F186 == True:
            self.Ter.AS_lineEdit.setEnabled(False)
        else:
            self.Ter.AS_lineEdit.setEnabled(True)
            
    
    def UpdateJSON(self):
        init_value = {  
            "Tab"           : {'Info_tabWidget':{}},
            "TabInfo"       : {"Testware" :{}, "Report" :{}},
            "TabSRCheck"    : {"SR_Filepath" :{}, "Control" :{}},
            "TabFSI"        : {"FSI_Filepath" :{}}	
        }           
        # ANCHOR - UpdateJSON - TabInfo ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
        # Tab:
        init_value['Tab']['Info_tabWidget']['CurrentIndex']              = self.Ter.Info_tabWidget.currentIndex()
        # Testware:
        init_value['TabInfo']['Testware']['WFP_lineEdit']                = self.Ter.WFP_lineEdit.text().strip().replace('\\','/')
        init_value['TabInfo']['Testware']['SWP_lineEdit']                = self.Ter.SWP_lineEdit.text().strip().replace('\\','/')
        init_value['TabInfo']['Testware']['PRJ_comboBox']                = self.Ter.PRJ_comboBox.currentText().strip()
        init_value['TabInfo']['Testware']['Var_comboBox']                = self.Ter.Var_comboBox.currentText().strip()
        init_value['TabInfo']['Testware']['Type_comboBox']               = self.Ter.Type_comboBox.currentText().strip()
        init_value['TabInfo']['Testware']['RCV_comboBox']                = self.Ter.RCV_comboBox.currentText().strip()
        init_value['TabInfo']['Testware']['PN_lineEdit']                 = self.Ter.PN_lineEdit.text().strip()
        init_value['TabInfo']['Testware']['BL_lineEdit']                 = self.Ter.BL_lineEdit.text().strip()
        init_value['TabInfo']['Testware']['RCV_lineEdit']                = self.Ter.RCV_lineEdit.text().strip()
        init_value['TabInfo']['Testware']['SWVer_lineEdit']              = self.Ter.SWVer_lineEdit.text().strip()
        init_value['TabInfo']['Testware']['Tester_comboBox']             = self.Ter.Tester_comboBox.currentText().strip()
        init_value['TabInfo']['Testware']['SWPCM_comboBox']              = self.Ter.SWPCM_comboBox.currentText().strip()
        init_value['TabInfo']['Testware']['TC_comboBox']                 = self.Ter.TC_comboBox.currentText().strip()
        init_value['TabInfo']['Testware']['TPM_lineEdit']                = self.Ter.TPM_lineEdit.text().strip()
        init_value['TabInfo']['Testware']['HWPCM_lineEdit']              = self.Ter.HWPCM_lineEdit.text().strip()
        init_value['TabInfo']['Testware']['Position']                    = self.Ter.Pos_comboBox.currentText().strip()
        # Report:
        init_value['TabInfo']['Report']['CheckList_Name']                = self.CheckList_Name
        init_value['TabInfo']['Report']['Protocol_Name']                 = self.Protocol_Name
        init_value['TabInfo']['Report']['Report_SS_lineEdit']            = self.Ter.Report_SS_lineEdit.text().strip().replace('\\','/')
        init_value['TabInfo']['Report']['Report_Protocol_lineEdit']      = self.Ter.Report_Protocol_lineEdit.text().strip().replace('\\','/')
        init_value['TabInfo']['Report']['Report_BKS_lineEdit']           = self.Ter.Report_BKS_lineEdit.text().strip().replace('\\','/')
        init_value['TabInfo']['Report']['Report_SRCR_lineEdit']          = self.Ter.Report_SRCR_lineEdit.text().strip().replace('\\','/')
        init_value['TabInfo']['Report']['BKS_lineEdit']                  = self.Ter.BKS_lineEdit.text().strip().replace('\\','/')
        init_value['TabInfo']['Report']['SS_lineEdit']                   = self.Ter.SS_lineEdit.text().strip().replace('\\','/')
        init_value['TabInfo']['Report']['SRCR_lineEdit']                 = self.Ter.SRCR_lineEdit.text().strip().replace('\\','/')
        init_value['TabInfo']['Report']['TTicket_lineEdit']              = self.Ter.TTicket_lineEdit.text().strip().replace('\\','/')
        init_value['TabInfo']['Report']['Reviewer_lineEdit']             = self.Ter.Reviewer_lineEdit.text().strip().replace('\\','/')
        init_value['TabInfo']['Report']['TM_lineEdit']                   = self.Ter.TM_lineEdit.text().strip().replace('\\','/')
        # ANCHOR - UpdateJSON - TabSRCheck ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
        # SR_Filepath:
        init_value['TabSRCheck']['SR_Filepath']['SB_lineEdit']           = self.Ter.SB_lineEdit.text().strip().replace('\\','/')
        init_value['TabSRCheck']['SR_Filepath']['Diatester_lineEdit']    = self.Ter.Diatester_lineEdit.text().strip().replace('\\','/')
        init_value['TabSRCheck']['SR_Filepath']['CANalyzer_lineEdit']    = self.Ter.CANalyzer_lineEdit.text().strip().replace('\\','/')
        init_value['TabSRCheck']['SR_Filepath']['CANape_lineEdit']       = self.Ter.CANape_lineEdit.text().strip().replace('\\','/')
        init_value['TabSRCheck']['SR_Filepath']['FaultID_lineEdit']      = self.Ter.FaultID_lineEdit.text().strip().replace('\\','/')
        init_value['TabSRCheck']['SR_Filepath']['PubDBC_lineEdit']       = self.Ter.PubDBC_lineEdit.text().strip().replace('\\','/')
        init_value['TabSRCheck']['SR_Filepath']['PriDBC_lineEdit']       = self.Ter.PriDBC_lineEdit.text().strip().replace('\\','/')
        init_value['TabSRCheck']['SR_Filepath']['Oxflash_lineEdit']      = self.Ter.Oxflash_lineEdit.text().strip().replace('\\','/')
        init_value['TabSRCheck']['SR_Filepath']['Nxflash_lineEdit']      = self.Ter.Nxflash_lineEdit.text().strip().replace('\\','/')
        init_value['TabSRCheck']['SR_Filepath']['ESBBOX_lineEdit']       = self.Ter.ESBBOX_lineEdit.text().strip().replace('\\','/')
        init_value['TabSRCheck']['SR_Filepath']['STIL_lineEdit']         = self.Ter.STIL_lineEdit.text().strip().replace('\\','/')
        # Control:
        init_value['TabSRCheck']['Control']['TXID_lineEdit']             = self.Ter.TXID_lineEdit.text().strip()
        init_value['TabSRCheck']['Control']['RXID_lineEdit']             = self.Ter.RXID_lineEdit.text().strip()
        init_value['TabSRCheck']['Control']['AS_lineEdit']               = self.Ter.AS_lineEdit.text().strip()
        init_value['TabSRCheck']['Control']['Diagnostic_lineEdit']       = self.Ter.Diagnostic_lineEdit.text().strip()
        init_value['TabSRCheck']['Control']['UDS_lineEdit']              = self.Ter.UDS_lineEdit.text().strip().replace('\\','/')
        init_value['TabSRCheck']['Control']['UDSSheet_lineEdit']         = self.Ter.UDSSheet_lineEdit.text().strip()
        init_value['TabSRCheck']['Control']['AS_checkBox']               = self.is_ActSes_F186      
        init_value['TabSRCheck']['Control']['DS_checkBox']               = self.is_DS_check      
        init_value['TabSRCheck']['Control']['PS_checkBox']               = self.is_PS_check      
        init_value['TabSRCheck']['Control']['ES_checkBox']               = self.is_ES_check      
        init_value['TabSRCheck']['Control']['PhysDiag_checkBox']         = self.is_PhysDiag_check
        init_value['TabSRCheck']['Control']['FuncDiag_checkBox']         = self.is_FuncDiag_check
        init_value['TabSRCheck']['Control']['Sub01_checkBox']            = self.is_Sub01_check   
        init_value['TabSRCheck']['Control']['Sub02_checkBox']            = self.is_Sub02_check   
        init_value['TabSRCheck']['Control']['Sub03_checkBox']            = self.is_Sub03_check   
        init_value['TabSRCheck']['Control']['SPRMIB_checkBox']           = self.is_SPRMIB_check  
        # ANCHOR - UpdateJSON - TabFSI -------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
        ## FSI_Filepath:
        init_value['TabFSI']['FSI_Filepath']['FSI_SB_lineEdit']          = self.Ter.FSI_SB_lineEdit.text().strip().replace('\\','/')
        init_value['TabFSI']['FSI_Filepath']['FSI_Diatester_lineEdit']   = self.Ter.FSI_Diatester_lineEdit.text().strip().replace('\\','/')
        init_value['TabFSI']['FSI_Filepath']['FSI_CANalyzer_lineEdit']   = self.Ter.FSI_CANalyzer_lineEdit.text().strip().replace('\\','/')
        init_value['TabFSI']['FSI_Filepath']['FSI_CANape_lineEdit']      = self.Ter.FSI_CANape_lineEdit.text().strip().replace('\\','/')
        init_value['TabFSI']['FSI_Filepath']['FSI_FaultID_lineEdit']     = self.Ter.FSI_FaultID_lineEdit.text().strip().replace('\\','/')
        init_value['TabFSI']['FSI_Filepath']['FSI_PubDBC_lineEdit']      = self.Ter.FSI_PubDBC_lineEdit.text().strip().replace('\\','/')
        init_value['TabFSI']['FSI_Filepath']['FSI_PriDBC_lineEdit']      = self.Ter.FSI_PriDBC_lineEdit.text().strip().replace('\\','/')
        init_value['TabFSI']['FSI_Filepath']['ClaraTE_lineEdit']         = self.Ter.ClaraTE_lineEdit.text().strip().replace('\\','/')
        
        with open(self.LoadInit_path, 'w') as latest_data:
            json.dump(init_value, latest_data, indent=4)

    # ------------------------#ANCHOR: Tab widget ------------------------:
    def Run_TabChanged_Func(self):
        SRPartner.UpdateJSON(self)

    # -------------------- #ANCHOR: Create database --------------------:
    def createDatabase(self):
        if 'SRPartnerTool' not in os.listdir(os.path.dirname(self.DB_dirname)):
            os.mkdir(os.path.join(self.temp,'SRPartnerTool'))
            shutil.copy(self.DB_org, self.DB_dirname)
            SRPartner.Console(self,'Pass', 'Database created')
        else:
            SRPartner.Console(self,'Status', 'Database already created')

    # -------------------- #ANCHOR: Progress --------------------:
    # ANCHOR - Compute
    def Compute(self, step, type, total):
        percent = int((step/total)*100)
        return percent

    # ANCHOR - Progress
    def Progress(self, step, type, total):
        percent = self.Compute(step, type, total)
        self.Ter.progressBar.setValue(percent)
        
    # -------------------- #ANCHOR: Tool-bar --------------------:
    # ANCHOR - Infomation
    def actionAbout(self):
        msg = QMessageBox()
        msg.setWindowTitle('''About SRPartner - FSI SRCheck support tool''')
        msg.setText('''About SRPartner:\n
        Author  : Nguyen Thanh Nhan
        NT-User : HNG4HC
        Version : 1.0''')
        msg.setIcon(QMessageBox.Information)
        msg.setStandardButtons(QMessageBox.Cancel)
        x = msg.exec_()
        
    def actionConfiguration(self):
        widget = QDialog(self)
        ui=Configuration_ui()
        ui.setupUi(widget)
        widget.exec_()
        
    def Config_SWVer(self):
        #BL-RC-V:  
        try:
            if int(self.Ter.BL_lineEdit.text()) < 10:
                if float(self.Ter.RCV_lineEdit.text()) < 10 :
                    BL  = '0{0}'.format(self.Ter.BL_lineEdit.text())
                    RC  = '0{0}'.format(self.Ter.RCV_lineEdit.text())
                    RCV = self.Ter.RCV_comboBox.currentText()
                    self.SWVer = 'BL{0}_{2}{1}'.format(BL,RC,RCV)
                else:
                    BL = '0{0}'.format(self.Ter.BL_lineEdit.text())
                    RC = '{0}'.format(self.Ter.RCV_lineEdit.text())
                    RCV = self.Ter.RCV_comboBox.currentText()
                    self.SWVer = 'BL{0}_{2}{1}'.format(BL,RC,RCV)
            else:
                if float(self.Ter.RCV_lineEdit.text()) < 10 :
                    BL = '{0}'.format(self.Ter.BL_lineEdit.text())
                    RC = '0{0}'.format(self.Ter.RCV_lineEdit.text())
                    RCV = self.Ter.RCV_comboBox.currentText()
                    self.SWVer = 'BL{0}_{2}{1}'.format(BL,RC,RCV)
                else:
                    BL = '{0}'.format(self.Ter.BL_lineEdit.text())
                    RC = '{0}'.format(self.Ter.RCV_lineEdit.text())
                    RCV = self.Ter.RCV_comboBox.currentText()
                    self.SWVer = 'BL{0}_{2}{1}'.format(BL,RC,RCV)
        except ValueError:
            SRPartner.Console(self,'Error', 'RC/V can not be empty')
            self.SWVer = 'ERROR'

    # -------------------- #ANCHOR: TEST Func --------------------:
    def TEST(self):
        TXID  = int(self.Ter.TXID_lineEdit.text(),16)
        RXID  = int(self.Ter.RXID_lineEdit.text(),16)
        ISCANFD = True
        Func_INS.TEST(TXID, RXID, ISCANFD)
        
    # -------------------- #ANCHOR: HTML --------------------:
    def fill_PrjInfo(self, report_path, prj_, release_, tester_, case_, dt_):
        with open(report_path) as rp:
            soup = BeautifulSoup(rp, 'lxml')
            for p in soup.find_all('p'):
                #Fill:
                if 'Project' in str(p):
                    p.string.replace_with(f'Project: {str(prj_)}')
                elif 'Release' in str(p):
                    p.string.replace_with(f'Release: {str(release_)}')
                elif 'Tester' in str(p):
                    p.string.replace_with(f'Tester: {str(tester_)}')
                elif 'Case' in str(p):
                    p.string.replace_with(f'Case: {str(case_)}')
                else:
                    p.string.replace_with(f'Date/time: {str(dt_)}')
        # save the file again
        with open(report_path, "w") as outf:
            outf.write(str(soup))
    
    def HTML_Report(self, dict_, testcase_):
        num = 1
        dt_string = datetime.now().strftime("[%d-%m-%Y_%H.%M.%S]")
        with open(f"{self.Current_path}/template/Report/General_template.html") as inf:
            soup = BeautifulSoup(inf.read())
        for key, val in dict(dict_).items():
            new_tr = soup.new_tag("tr")
            #Add as many td (data) 
            No                = soup.new_tag('td',attrs={"class": "tdbreak"})
            TC_Step           = soup.new_tag('td',attrs={"class": "tdbreak"})
            Name              = soup.new_tag('td',attrs={"class": "tdbreak"})
            Request           = soup.new_tag('td',attrs={"class": "tdbreak"})
            Response          = soup.new_tag('td',attrs={"class": "tdbreak"})
            Ascii             = soup.new_tag('td',attrs={"class": "tdbreak"})
            Len_Check         = soup.new_tag('td',attrs={"class": "tdbreak"})
            Val_Check = soup.new_tag('td',attrs={"class": "tdbreak"})
            #Set value:
            No.string                = str(num)
            TC_Step.string           = str(val['TC_Step'])
            Name.string              = str(val['Name']) + ' ' + key
            Request.string           = str(val['Request']) 
            Response.string          = str(val['Response'])
            Ascii.string             = Func_INS.hex2ascii()
            Len_Check.string         = str(val['Length'])
            Val_Check.string = str(val['Expect_Val'])
            new_tr.append(No)
            new_tr.append(TC_Step)
            new_tr.append(did)
            new_tr.append(Name)
            new_tr.append(Request)
            new_tr.append(Response)
            new_tr.append(Len_Check)
            new_tr.append(Val_Check)
            #Add whole 'tr'(row) to table.
            soup.table.append(new_tr)
            num +=1
            
        # save the file again
        report_path = ''
        try: 
            report_path = f"{os.path.join(self.Ter.WFP_lineEdit.text(),'PIC', 'log', 'Test_log')}/Case_{testcase_}_{dt_string}.html"
            with open(report_path, "w") as outf:
                outf.write(str(soup))
        except: 
            report_path = f"{self.Current_path}/temp/report/Case_{testcase_}_{dt_string}.html"
            with open(report_path, "w") as outf:
                outf.write(str(soup))  
        # Fill project info:
        Prj      = str(self.LoadInit_dict['TabInfo']['Testware']['Var_comboBox'])
        release  = str(self.LoadInit_dict['TabInfo']['Testware']['SWVer_lineEdit'])
        tester   = str(self.LoadInit_dict['TabInfo']['Testware']['Tester_comboBox'])
        case     = str(testcase_) 
        dt       = str(dt_string)
        SRPartner.fill_PrjInfo(self,report_path,Prj,release,tester,case,dt)
    
    def case_7_1_HTML(self, dict_):
        num = 1
        inform = 'Update later'
        dt_string = datetime.now().strftime("[%d-%m-%Y_%H.%M.%S]")
        with open(f"{self.Current_path}/template/Report/Case_7_template.html") as inf:
            soup = BeautifulSoup(inf.read())
        for key, val in dict(dict_).items():
            length = len(val['Request'])
            for i in range(length):
                new_tr = soup.new_tag("tr")
                #Add as many td (data) 
                No       = soup.new_tag('td',attrs={"class": "tdbreak"})
                TC_name  = soup.new_tag('td',attrs={"class": "tdbreak"})
                Diag     = soup.new_tag('td',attrs={"class": "tdbreak"})
                Request  = soup.new_tag('td',attrs={"class": "tdbreak"})
                Response = soup.new_tag('td',attrs={"class": "tdbreak"})
                Check_1  = soup.new_tag('td',attrs={"class": "tdbreak"})
                Check_2  = soup.new_tag('td',attrs={"class": "tdbreak"})
                #Set value:
                No.string                = str(num)
                TC_name.string           = str(val['TC_name'])
                Diag.string              = str(val['Diag'])
                Request.string           = str(val['Request'][i])
                Response.string          = str(val['Response'][i])
                Check_1.string           = str(inform)
                Check_2.string           = str(inform)
                new_tr.append(No)
                new_tr.append(TC_name)
                new_tr.append(Diag)
                new_tr.append(Request)
                new_tr.append(Response)
                new_tr.append(Check_1)
                new_tr.append(Check_2)
                #Add whole 'tr'(row) to table.
                soup.table.append(new_tr) 
                num +=1 
        
        # save the file again
        report_path = ''
        try: 
            report_path = f"{os.path.join(self.Ter.WFP_lineEdit.text(),'PIC', 'log', 'Test_log')}/Case_7_1_{dt_string}.html"
            with open(report_path, "w") as outf:
                outf.write(str(soup))
        except: 
            report_path = f"{self.Current_path}/temp/report/Case_7_1_{dt_string}.html"
            with open(report_path, "w") as outf:
                outf.write(str(soup))  
        # Fill project info:
        Prj      = str(self.LoadInit_dict['TabInfo']['Testware']['Var_comboBox'])
        release  = str(self.LoadInit_dict['TabInfo']['Testware']['SWVer_lineEdit'])
        tester   = str(self.LoadInit_dict['TabInfo']['Testware']['Tester_comboBox'])
        case     = str(7.1)
        dt       = str(dt_string)
        SRPartner.fill_PrjInfo(self,report_path,Prj,release,tester,case,dt)
            
            
            
########## !SECTION: MainWindow  ##########

if __name__ == "__main__":
    app = QApplication(sys.argv) 
    main_win = SRPartner()
    main_win.show()
    sys.exit(app.exec())