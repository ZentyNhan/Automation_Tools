'''
+----------------------------------------------------------------------------+
License (C) Copyright 2022-2023,  Corporation Limited.
Author            :    Nguyen Thanh Nhan, Phan Manh Duc
Python Version    :    Python3.7.0
Email Address     :    Nhan.NguyenThanh5@vn.bosch.com, Duc.PhanManh@vn.bosch.com
Creat Time        :    Jul 8, 2022 10:10:10
Description       :    Billing tool
------------------------------------------------------------------------------
Modification History
     Data                  By             Version       Change Description
==============================================================================

==============================================================================
'''
from contextlib import nullcontext
from itertools import count
from re import X
from tkinter import W
from xml.dom.expatbuilder import ElementInfo

import openpyxl
from pytest import PytestCollectionWarning
import BillingTools
from sqlite3 import connect
from PyQt5 import QtGui, QtWidgets
from PyQt5.QtGui import QStandardItemModel
from PyQt5.QtGui import QStandardItem
from PyQt5.QtWidgets import QMainWindow, QApplication , QDialog, QTabWidget,QDateEdit, QHBoxLayout, QGridLayout, QGroupBox, QVBoxLayout, QWidget, QLabel, QLineEdit, QDialogButtonBox, QMessageBox, QPushButton, QFileDialog, QTextEdit, QListWidget, QAbstractItemView, QComboBox
from PyQt5.QtCore import QThread, pyqtSignal, QObject, QDate, QTime, QDateTime
from PyQt5.QtWidgets import QCalendarWidget
from PyQt5.uic import loadUi
import threading
import sys
import json
import time
import datetime
import os, webbrowser
import socket
from PyQt5 import QtTest
from PyQt5 import QtCore
from PyQt5.QtCore import QThread
from PyQt5.QtCore import QTimer
from BillingTools import Ui_BillingTool
from Lib.Template_init import Template_UI
from PyQt5.QtWidgets import QDateEdit
from termcolor import colored 
from datetime import datetime
from datetime import date
import win32com.client.gencache as gencache
from Lib.EmployeeDict import EmployeeDict
#Excel Processing:
from openpyxl import Workbook 
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font,PatternFill,Border,Side
from openpyxl.utils.exceptions import InvalidFileException

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_BillingTool()
        self.ui.setupUi(self)

        # ANCHOR: Init variables
        ##-----Flags-----##:

        ##-----Variables-----##:
        self.Allowcate_link = nullcontext
        self.AssociateList_link = nullcontext
        self.ProjectList_link = nullcontext
        self.Report_link = nullcontext
        self.WB_Project = Workbook()
        self.WSN_Allocate_G4 = nullcontext
        self.WSN_Allocate_G5 = nullcontext
        self.ProjectDict = {}

        #Date From:
        self.Day_DateFrom = int(0)
        self.Month_DateFrom = int(0)
        self.Year_DateFrom = int(0)
        self.Format_DateFrom = int(0)
        self.Selected_DateFrom = int(0)

        #Date To:
        self.Day_DateTo = int(0)
        self.Month_DateTo = int(0)
        self.Year_DateTo = int(0)
        self.Format_DateTo = int(0)
        self.Selected_DateTo = int(0)

        #TEST:
        self.TEST = nullcontext

        #Path:
        # Get Current path:
        self.Current_path = os.path.dirname(__file__)
        # self.Current_path = self.Current_path.replace('\\','/')
        self.Current_string = nullcontext
       
        #Reports:
        self.WB_Report = openpyxl.load_workbook(filename='{0}/Template/Excel/EDA3_SystemTestTeam_Billing_Template.xlsx'.format(self.Current_path), read_only=False, keep_vba=False,
                    data_only=False, keep_links=True)

        self.Report_Link_save = nullcontext
        
        #ANCHOR - Events
        #---Buttton---:
        self.ui.Browse_Button.clicked.connect(self.OpenBrowse_Allowcate_Function)
        self.ui.Projects_Browse_Button.clicked.connect(self.Project_Browse_Function)
        self.ui.Report_Browse_Button.clicked.connect(self.Report_Browse_Function)
        self.ui.Execute_pushButton.clicked.connect(self.Execute_Billing_Function)
        self.ui.OpenReport_pushButton.clicked.connect(self.Open_Report_Function)
        self.ui.Extract_Button.clicked.connect(self.Extract_Data_Function)
        self.ui.Re_Calculate_Button.clicked.connect(self.Re_Calculate_Report_Function)
        self.ui.OpenFolderReport_pushButton.clicked.connect(self.OpenFolderReport_pushButton_Function)
        self.ui.Browse_Clear_pushButton.clicked.connect(self.Clear_Allocate_Function)
        self.ui.Projects_Browse_Clear_pushButton.clicked.connect(self.Clear_Project_Function)
        self.ui.Report_Browse_Clear_pushButton.clicked.connect(self.Clear_Report_Function)
        #---Widget---:
        self.ui.DateFrom_calendarWidget.selectionChanged.connect(self.CalendarChanged_DateFrom_Function)
        self.ui.DateFrom_calendarWidget.setObjectName("DateFrom_calendarWidget")
        self.ui.DateFrom_calendarWidget.show()
        self.ui.DateTo_calendarWidget.selectionChanged.connect(self.CalendarChanged_DateTo_Function) 
        self.ui.DateTo_calendarWidget.setObjectName("DateTo_calendarWidget")
        self.ui.DateTo_calendarWidget.hide() 
        #---LineEdit---:
        self.ui.Allocate_lineEdit.textEdited.connect(self.Allocate_lineEdit_Function)
        self.ui.Projects_lineEdit.textEdited.connect(self.Projects_lineEdit_Function)
        self.ui.Allocate_G4_Sheet_lineEdit.textEdited.connect(self.Allocate_G4_Sheet_lineEdit_Function)
        self.ui.Allocate_G5_Sheet_lineEdit.textEdited.connect(self.Allocate_G5_Sheet_lineEdit_Function)
        self.ui.Projects_Sheet_lineEdit.textEdited.connect(self.Projects_Sheet_lineEdit_Function)
        #---Combobox---:
        #Add items:
        Items = ['......','Date From','Date To']
        self.ui.Date_comboBox.addItems(Items)
        self.ui.Date_comboBox.currentTextChanged.connect(self.Choose_DateStatus_Function)

        ################## TEST MODE ##################
        self.ui.Test_button.clicked.connect(self.DO_SOMETHING_TEST)
        self.ui.Allocate_lineEdit.textEdited.connect(self.DO_SOMETHING_TEST_1)
        ################## TEST MODE ##################

    ##-----Functions-----##:
    #------------------------Browse Button------------------------#
    def OpenBrowse_Allowcate_Function(self):
        try:
            file_filter = 'All Files (*.*)'
            OpenObject = QFileDialog.getOpenFileName(parent= self, caption="Open Allocate file", directory='{0}/Template/Excel'.format(self.Current_path), filter=file_filter )
            self.ui.Allocate_lineEdit.setText(OpenObject[0])
            self.Allowcate_link = self.ui.Allocate_lineEdit.text()  
            self.WB_Allocate = openpyxl.load_workbook(filename= (self.Allowcate_link), read_only=False, keep_vba=False,
                  data_only=False, keep_links=True)
            if self.ui.Allocate_G4_Sheet_lineEdit.text() == '' or self.ui.Allocate_G5_Sheet_lineEdit.text() == '':
                self.ui.Status_Console_textBrowser.setStyleSheet("color: rgb(255, 0, 0);") #set text color
                self.ui.Status_Console_textBrowser.setText('Please insert Allocate sheet name')
            elif self.ui.Allocate_G4_Sheet_lineEdit.text() == self.ui.Allocate_G5_Sheet_lineEdit.text():
                self.ui.Status_Console_textBrowser.setStyleSheet("color: rgb(255, 0, 0);") #set text color
                self.ui.Status_Console_textBrowser.setText('Sheet Gen4 and Gen5 should be different')
            else:
                self.WSN_Allocate_G4 = self.ui.Allocate_G4_Sheet_lineEdit.text()
                self.WSN_Allocate_G5 = self.ui.Allocate_G5_Sheet_lineEdit.text()
                self.WB_Allocate[self.WSN_Allocate_G4]
                self.WB_Allocate[self.WSN_Allocate_G5]
                self.ui.Status_Console_textBrowser.setStyleSheet("font-weight: bold; color: green") #set text color
                self.ui.Status_Console_textBrowser.setText('sheet name is ok!')
        except KeyError:
            self.ui.Status_Console_textBrowser.setStyleSheet("font-weight: bold; color: red") #set text color
            self.ui.Status_Console_textBrowser.setText('Wrong Sheet name. Please try again')
        except InvalidFileException:
            pass


    def Project_Browse_Function(self):
        try:
            file_filter = 'All Files (*.*)' 
            OpenObject = QFileDialog.getOpenFileName(parent= self, caption="Open Projects file", directory='{0}/Template/Excel'.format(self.Current_path), filter=file_filter )
            self.ui.Projects_lineEdit.setText(OpenObject[0])
            self.ProjectList_link = self.ui.Projects_lineEdit.text()  
            self.WB_Project = openpyxl.load_workbook(filename= (self.ProjectList_link), read_only=False, keep_vba=False,
                  data_only=False, keep_links=True)
            if self.ui.Projects_Sheet_lineEdit.text() == '':
                self.ui.Status_Console_textBrowser.setStyleSheet("color: rgb(255, 0, 0);") #set text color
                self.ui.Status_Console_textBrowser.setText('Please insert Projects sheet name')
            else:
                self.WS_Project = self.WB_Project[self.ui.Projects_Sheet_lineEdit.text()]
                self.ProjectDict = Template_UI.Project_Dict(self.WS_Project)
                self.ui.Status_Console_textBrowser.setStyleSheet("font-weight: bold; color: green") #set text color
                self.ui.Status_Console_textBrowser.setText('sheet name is ok!')
        except KeyError:
            self.ui.Status_Console_textBrowser.setStyleSheet("font-weight: bold; color: red") #set text color
            self.ui.Status_Console_textBrowser.setText('Wrong Sheet name. Please try again')
        except InvalidFileException:
            pass

    def Report_Browse_Function(self):
        try:
            file_filter = 'All Files (*.*)' 
            OpenObject = QFileDialog.getOpenFileName(parent= self, caption="Open report file", directory='{0}/Report'.format(self.Current_path), filter=file_filter )
            self.ui.Report_lineEdit.setText(OpenObject[0])
            self.Report_link = self.ui.Report_lineEdit.text()
        except InvalidFileException:
            pass  

    #------------------------lineEdit------------------------#
    #Fill Allocate by manual:
    def Allocate_lineEdit_Function(self):
        try: 
            self.Allowcate_link = self.ui.Allocate_lineEdit.text()  
            self.WB_Allocate = openpyxl.load_workbook(filename= (self.Allowcate_link), read_only=False, keep_vba=False,
                  data_only=False, keep_links=True)
            if self.ui.Allocate_G4_Sheet_lineEdit.text() == '' or self.ui.Allocate_G5_Sheet_lineEdit.text() == '':
                self.ui.Status_Console_textBrowser.setStyleSheet("color: rgb(255, 0, 0);") #set text color
                self.ui.Status_Console_textBrowser.setText('Please insert Allocate sheet name')
            elif self.ui.Allocate_G4_Sheet_lineEdit.text() == self.ui.Allocate_G5_Sheet_lineEdit.text():
                self.ui.Status_Console_textBrowser.setStyleSheet("color: rgb(255, 0, 0);") #set text color
                self.ui.Status_Console_textBrowser.setText('Sheet Gen4 and Gen5 should be different')
            else:
                self.WSN_Allocate_G4 = self.ui.Allocate_G4_Sheet_lineEdit.text()
                self.WSN_Allocate_G5 = self.ui.Allocate_G5_Sheet_lineEdit.text()
                self.WB_Allocate[self.WSN_Allocate_G4]
                self.WB_Allocate[self.WSN_Allocate_G5]
                self.ui.Status_Console_textBrowser.setStyleSheet("font-weight: bold; color: green") #set text color
                self.ui.Status_Console_textBrowser.setText('sheet name is ok!')
        except KeyError:
            self.ui.Status_Console_textBrowser.setStyleSheet("font-weight: bold; color: red") #set text color
            self.ui.Status_Console_textBrowser.setText('Wrong Sheet name. Please try again')
        except InvalidFileException:
            pass  

    #Fill Projects by manual:
    def Projects_lineEdit_Function(self):
        try: 
            self.ProjectList_link = self.ui.Projects_lineEdit.text()  
            self.WB_Project = openpyxl.load_workbook(filename= (self.ProjectList_link), read_only=False, keep_vba=False,
                data_only=False, keep_links=True)
            if self.ui.Projects_Sheet_lineEdit.text() == '':
                self.ui.Status_Console_textBrowser.setStyleSheet("color: rgb(255, 0, 0);") #set text color
                self.ui.Status_Console_textBrowser.setText('Please insert Projects sheet name')
            else:
                self.WS_Project = self.WB_Project[self.ui.Projects_Sheet_lineEdit.text()]
                self.ProjectDict = Template_UI.Project_Dict(self.WS_Project)
                self.ui.Status_Console_textBrowser.setStyleSheet("font-weight: bold; color: green") #set text color
                self.ui.Status_Console_textBrowser.setText('sheet name is ok!')
        except KeyError:
            self.ui.Status_Console_textBrowser.setStyleSheet("font-weight: bold; color: red") #set text color
            self.ui.Status_Console_textBrowser.setText('Wrong Sheet name. Please try again')
        except InvalidFileException:
            pass  

    def Allocate_G4_Sheet_lineEdit_Function(self):
        try: 
            self.Allowcate_link = self.ui.Allocate_lineEdit.text()  
            self.WB_Allocate = openpyxl.load_workbook(filename= (self.Allowcate_link), read_only=False, keep_vba=False,
                data_only=False, keep_links=True)
            if self.ui.Allocate_G4_Sheet_lineEdit.text() == '' or self.ui.Allocate_G5_Sheet_lineEdit.text() == '':
                self.ui.Status_Console_textBrowser.setStyleSheet("color: rgb(255, 0, 0);") #set text color
                self.ui.Status_Console_textBrowser.setText('Please insert Allocate sheet name')
            elif self.ui.Allocate_G4_Sheet_lineEdit.text() == self.ui.Allocate_G5_Sheet_lineEdit.text():
                self.ui.Status_Console_textBrowser.setStyleSheet("color: rgb(255, 0, 0);") #set text color
                self.ui.Status_Console_textBrowser.setText('Sheet Gen4 and Gen5 should be different')
            else:
                self.WSN_Allocate_G4 = self.ui.Allocate_G4_Sheet_lineEdit.text()
                self.WSN_Allocate_G5 = self.ui.Allocate_G5_Sheet_lineEdit.text()
                self.WB_Allocate[self.WSN_Allocate_G4]
                self.WB_Allocate[self.WSN_Allocate_G5]
                self.ui.Status_Console_textBrowser.setStyleSheet("font-weight: bold; color: green") #set text color
                self.ui.Status_Console_textBrowser.setText('sheet name is ok!')
        except KeyError:
            self.ui.Status_Console_textBrowser.setStyleSheet("font-weight: bold; color: red") #set text color
            self.ui.Status_Console_textBrowser.setText('Wrong Sheet name. Please try again')
        except InvalidFileException:
            pass

    def Allocate_G5_Sheet_lineEdit_Function(self):
        try: 
            self.Allowcate_link = self.ui.Allocate_lineEdit.text()  
            self.WB_Allocate = openpyxl.load_workbook(filename= (self.Allowcate_link), read_only=False, keep_vba=False,
                data_only=False, keep_links=True)
            if self.ui.Allocate_G4_Sheet_lineEdit.text() == '' or self.ui.Allocate_G5_Sheet_lineEdit.text() == '':
                self.ui.Status_Console_textBrowser.setStyleSheet("color: rgb(255, 0, 0);") #set text color
                self.ui.Status_Console_textBrowser.setText('Please insert Allocate sheet name')
            elif self.ui.Allocate_G4_Sheet_lineEdit.text() == self.ui.Allocate_G5_Sheet_lineEdit.text():
                self.ui.Status_Console_textBrowser.setStyleSheet("color: rgb(255, 0, 0);") #set text color
                self.ui.Status_Console_textBrowser.setText('Sheet Gen4 and Gen5 should be different')
            else:
                self.WSN_Allocate_G4 = self.ui.Allocate_G4_Sheet_lineEdit.text()
                self.WSN_Allocate_G5 = self.ui.Allocate_G5_Sheet_lineEdit.text()
                self.WB_Allocate[self.WSN_Allocate_G4]
                self.WB_Allocate[self.WSN_Allocate_G5]
                self.ui.Status_Console_textBrowser.setStyleSheet("font-weight: bold; color: green") #set text color
                self.ui.Status_Console_textBrowser.setText('sheet name is ok!')
        except KeyError:
            self.ui.Status_Console_textBrowser.setStyleSheet("font-weight: bold; color: red") #set text color
            self.ui.Status_Console_textBrowser.setText('Wrong Sheet name. Please try again')
        except InvalidFileException:
            pass
        
    def Projects_Sheet_lineEdit_Function(self):
        try: 
            self.ProjectList_link = self.ui.Projects_lineEdit.text()  
            self.WB_Project = openpyxl.load_workbook(filename= (self.ProjectList_link), read_only=False, keep_vba=False,
                data_only=False, keep_links=True)
            if self.ui.Projects_Sheet_lineEdit.text() == '':
                self.ui.Status_Console_textBrowser.setStyleSheet("color: rgb(255, 0, 0);") #set text color
                self.ui.Status_Console_textBrowser.setText('Please insert Projects sheet name')
            else:
                self.WS_Project = self.WB_Project[self.ui.Projects_Sheet_lineEdit.text()]
                self.ProjectDict = Template_UI.Project_Dict(self.WS_Project)
                self.ui.Status_Console_textBrowser.setStyleSheet("font-weight: bold; color: green") #set text color
                self.ui.Status_Console_textBrowser.setText('sheet name is ok!')
        except KeyError:
            self.ui.Status_Console_textBrowser.setStyleSheet("font-weight: bold; color: red") #set text color
            self.ui.Status_Console_textBrowser.setText('Wrong Sheet name. Please try again')
        except InvalidFileException:
            pass

    #------------------------Widget------------------------#
    def CalendarChanged_DateFrom_Function(self):
            #Save value:
            self.Day_DateFrom = self.ui.DateFrom_calendarWidget.selectedDate().day()
            self.Month_DateFrom = self.ui.DateFrom_calendarWidget.selectedDate().month()
            self.Year_DateFrom = self.ui.DateFrom_calendarWidget.selectedDate().year()
            #Format Arrange:
            self.Format_DateFrom =  '{0}/{1}/{2}'.format(self.Year_DateFrom, self.Month_DateFrom, self.Day_DateFrom)
            #Display In TestBrowser:
            self.ui.DateFrom_textBrowser.setText(self.Format_DateFrom)
            self.Selected_DateFrom = date(self.Year_DateFrom, self.Month_DateFrom, self.Day_DateFrom)

    def CalendarChanged_DateTo_Function(self):
            #Save value:
            self.Day_DateTo = self.ui.DateTo_calendarWidget.selectedDate().day()
            self.Month_DateTo = self.ui.DateTo_calendarWidget.selectedDate().month()
            self.Year_DateTo = self.ui.DateTo_calendarWidget.selectedDate().year()
            #Format Arrange: 
            self.Format_DateTo =  '{0}/{1}/{2}'.format(self.Year_DateTo, self.Month_DateTo, self.Day_DateTo)
            #Display In TestBrowser:
            self.ui.DateTo_textBrowser.setText(self.Format_DateTo)
            self.Selected_DateTo = date(self.Year_DateTo, self.Month_DateTo, self.Day_DateTo)

    #------------------------Clear text------------------------#
    def Clear_Allocate_Function(self):
        self.ui.Allocate_lineEdit.clear()
        self.ui.Status_Console_textBrowser.clear()

    def Clear_Project_Function(self):
        self.ui.Projects_lineEdit.clear()
        self.ui.Status_Console_textBrowser.clear()

    def Clear_Report_Function(self):
        self.ui.Report_lineEdit.clear()
        self.ui.Status_Console_textBrowser.clear()

    #------------------------Combobox------------------------#
    #ANCHOR - Choose_DateStatus_Function
    def Choose_DateStatus_Function(self):
        if self.ui.Date_comboBox.currentText() in ['Date From']:   
            #Enable calendar:
            self.ui.DateFrom_calendarWidget.setEnabled(True)
            self.ui.DateTo_calendarWidget.setEnabled(True) 
            #Show Calendar:
            self.ui.DateFrom_calendarWidget.show()
            self.ui.DateTo_calendarWidget.hide()
            #Set Font:
            font_DateFrom = QtGui.QFont()
            font_DateFrom.setPointSize(13)
            font_DateFrom.setBold(True)
            font_DateFrom.setCapitalization(1)
            self.ui.DateFrom_Label.setFont(font_DateFrom)
            font_DateTo = QtGui.QFont()
            font_DateTo.setPointSize(12)
            font_DateTo.setBold(False)
            self.ui.DateTo_Label.setFont(font_DateTo)

        elif self.ui.Date_comboBox.currentText() in ['Date To']:
            #Enable calendar:
            self.ui.DateFrom_calendarWidget.setEnabled(True)
            self.ui.DateTo_calendarWidget.setEnabled(True)
            #Show Calendar:        
            self.ui.DateFrom_calendarWidget.hide()
            self.ui.DateTo_calendarWidget.show() 
            #Set Font:
            font_DateFrom = QtGui.QFont()
            font_DateFrom.setPointSize(12)
            font_DateFrom.setBold(False)
            self.ui.DateFrom_Label.setFont(font_DateFrom)
            font_DateTo = QtGui.QFont()
            font_DateTo.setPointSize(13)
            font_DateTo.setBold(True)
            font_DateTo.setCapitalization(1)
            self.ui.DateTo_Label.setFont(font_DateTo)
        else:
            self.ui.DateFrom_calendarWidget.setEnabled(False)
            self.ui.DateTo_calendarWidget.setEnabled(False)
            #Set Font:
            font_DateFrom = QtGui.QFont()
            font_DateFrom.setPointSize(12)
            font_DateFrom.setBold(False)
            self.ui.DateFrom_Label.setFont(font_DateFrom)
            font_DateTo = QtGui.QFont()
            font_DateTo.setPointSize(12)
            font_DateTo.setBold(False)
            self.ui.DateTo_Label.setFont(font_DateTo)
    
    #------------------------Execute Function------------------------#
    #########################
    ### STAGE 1: Get data ###
    #########################
    # ANCHOR: Extract data: 
    def Extract_Data_Function(self):
        #re-assign
        self.Allowcate_link = self.ui.Allocate_lineEdit.text()
        self.ProjectList_link = self.ui.Projects_lineEdit.text()
        #Check condition:
        if self.Allowcate_link == nullcontext or len(self.Allowcate_link) == 0:
            self.ui.Status_Console_textBrowser.setStyleSheet("color: rgb(255, 0, 0);") #set text color
            self.ui.Status_Console_textBrowser.setText('Please insert Allocate file')
        elif self.ProjectList_link == nullcontext or len(self.ProjectList_link) == 0:
            self.ui.Status_Console_textBrowser.setStyleSheet("color: rgb(255, 0, 0);") #set text color
            self.ui.Status_Console_textBrowser.setText('Please insert Projects list file')
        elif self.ui.Allocate_G4_Sheet_lineEdit.text() == nullcontext or \
                self.ui.Allocate_G5_Sheet_lineEdit.text() == nullcontext or \
                self.ui.Projects_Sheet_lineEdit.text() == nullcontext:
            self.ui.Status_Console_textBrowser.setStyleSheet("color: rgb(255, 0, 0);") #set text color
            self.ui.Status_Console_textBrowser.setText('Please insert sheet name')
        elif self.ui.Allocate_G4_Sheet_lineEdit.text() == self.ui.Allocate_G5_Sheet_lineEdit.text():
                self.ui.Status_Console_textBrowser.setStyleSheet("color: rgb(255, 0, 0);") #set text color
                self.ui.Status_Console_textBrowser.setText('Sheet Gen4 and Gen5 should be different')
        elif (self.Format_DateFrom == int(0)) or (self.Format_DateTo == int(0)):
            self.ui.Status_Console_textBrowser.setStyleSheet("color: rgb(255, 0, 0);") #set text color
            self.ui.Status_Console_textBrowser.setText('Please insert date')
        else:
            if self.Year_DateFrom > self.Year_DateTo:
                self.ui.Status_Console_textBrowser.setStyleSheet("color: rgb(255, 0, 0);")
                self.ui.Status_Console_textBrowser.setText('INVALID INSERT VALUE: Wrong YEAR value!\nPlease try again')
            elif (self.Year_DateFrom <= self.Year_DateTo) and (self.Month_DateFrom > self.Month_DateTo):
                self.ui.Status_Console_textBrowser.setStyleSheet("color: rgb(255, 0, 0);") #set text color
                self.ui.Status_Console_textBrowser.setText('INVALID INSERT VALUE: Wrong MONTH value!\nPlease try again')
            elif (self.Year_DateFrom == self.Year_DateTo) and (self.Month_DateFrom == self.Month_DateTo) and (self.Day_DateFrom > self.Day_DateTo):
                self.ui.Status_Console_textBrowser.setStyleSheet("color: rgb(255, 0, 0);") #set text color
                self.ui.Status_Console_textBrowser.setText('INVALID INSERT VALUE: Wrong DAY value!\nPlease try again')
            else: 
                self.ui.Status_Console_textBrowser.clear()

                #################################
                #FIXME : ĐIỀN MERGE CODE VÔ ĐÂY:#
                #################################
                #Local variables:
                now = datetime.now()
                dt_string = now.strftime("%d-%m-%Y_%H.%M.%S")
                AllocateResource_ws = self.WB_Allocate['Gen5_2022']
                AllocateResource_ws_4 = self.WB_Allocate[self.WSN_Allocate_G4]
                AllocateResource_ws_5 = self.WB_Allocate[self.WSN_Allocate_G5]
                Gen_4 = self.WB_Report.create_sheet('Gen_4')
                Gen_5 = self.WB_Report.create_sheet('Gen_5')
                OT_Sheet = self.WB_Report.create_sheet('OT')
                str_ignore_temp = [	
                        "COMPETENCY","NONE","LEAVE","HOLIDAY","RUN PROJECT WITH DUY",
                        "GEN4","LEAVE 1ST HALF","EXIT","GEN5 SR CHECK","CLARA","UT","X","TC",
                        "TRAINING","VTESTSTUDIO","KO NHẬN OT","UPS MAINTAINANCE IN AFTERNOON 14:00 - 17:00",
                        "GEN4 BYD SR","FSI GEN5","Support Huy","Support Nhat","SUPPORT VIDEO","Gen5","Afternoon leave",
                        " AUDI","LEAVE 2ND HALF","SUPPORT VIDEO","Dept Outing","GEN4 SRCheck","Gen4JMC","Phu Support Hung","Dept. Outing",
                        "Training SR_Check", "Support Video","AUDI","LAB Support","Support Build Lab","GEN4 JMC","SAIC","Video","LAB Support"
                    ]
                
                date_start 	= datetime(self.Year_DateFrom, self.Month_DateFrom, self.Day_DateFrom)
                date_end 	= datetime(self.Year_DateTo, self.Month_DateTo, self.Day_DateTo)

                str_ignore = []
                for x in str_ignore_temp:
                    str_ignore.append(x.upper())
                col_start	= ""
                col_end 	= ""
                for i in range(1, AllocateResource_ws.max_column + 1):
                    char =  get_column_letter(i)
                    if AllocateResource_ws[char + str(1)].value == date_start:
                        col_start = i
                    if AllocateResource_ws[char + str(1)].value == date_end:
                        col_end = i

                #Get OT data:
                OT_Date = Template_UI.Date_OT(AllocateResource_ws)

                headings = ['Customer'] + ['Project'] + ['BM_Number'] + ['Task_ID'] + ['Name_Employee'] + ['NumberID'] + ['Task'] + ['Estimate Date For Work']
                Gen_5.append(headings)
                Gen_4.append(headings)

                ot_headings = ['Customer'] + ['Project'] + ['BM_Number'] + ['Task_ID'] + ['Name_Employee'] + ['NumberID'] + ['Task'] + ['Date']
                OT_Sheet.append(ot_headings)

                for i in range(1, 80):
                    char_5 = get_column_letter(2)
                    name_employee_5 = str(AllocateResource_ws_5[char_5 + str(i)].value).replace("\n"," ").upper()
                    for Employee in EmployeeDict.values():
                        str_name_ee = Employee['Name'].split()
                        same_name = [s for s in str_name_ee if s not in (set(str_name_ee) ^ set(name_employee_5.split()))]
                        if (same_name == str_name_ee):
                            Template_UI.Gen_5_to_xlxs(i,EmployeeDict[Employee['Name']]['Name'],EmployeeDict[Employee['Name']]['NumberID'],
                                col_start, col_end, OT_Date, AllocateResource_ws_5, str_ignore, self.ProjectDict, Gen_5)
                            Template_UI.OT_to_xlxs_5(i,EmployeeDict[Employee['Name']]['Name'],EmployeeDict[Employee['Name']]['NumberID'], 
                                col_start, col_end, OT_Date, AllocateResource_ws_5, str_ignore, self.ProjectDict, OT_Sheet)
                        else:
                            continue

                for i in range(1, 65):
                    char_4 = get_column_letter(2)
                    name_employee_4 = str(AllocateResource_ws_4[char_4 + str(i)].value).replace("\n"," ").upper()
                    for Employee in EmployeeDict.values():
                        str_name_ee = Employee['Name'].split()
                        same_name = [s for s in str_name_ee if s not in (set(str_name_ee) ^ set(name_employee_4.split()))]
                        if (same_name == str_name_ee):
                            Template_UI.Gen_4_to_xlxs(i,EmployeeDict[Employee['Name']]['Name'],EmployeeDict[Employee['Name']]['NumberID'], 
                                col_start, col_end, OT_Date, AllocateResource_ws_4, str_ignore, self.ProjectDict, Gen_4)
                            Template_UI.OT_to_xlxs_4(i,EmployeeDict[Employee['Name']]['Name'],EmployeeDict[Employee['Name']]['NumberID'], 
                                col_start, col_end, OT_Date, AllocateResource_ws_4, str_ignore, self.ProjectDict, OT_Sheet)
                        else:
                            continue

                self.Current_string = dt_string
                #Check if 'Report' folder exist or not       
                if 'Report' not in os.listdir(self.Current_path):
                    #Create a 'Report' folder to contain reports
                    os.mkdir('Report') 

                    #Gen report:
                    self.Report_Link_save = '{0}/Report/Report_{1}.xlsx'.format(self.Current_path,self.Current_string)
                    self.WB_Report.save(self.Report_Link_save)
                else:
                    #Gen report:
                    self.Report_Link_save = '{0}/Report/Report_{1}.xlsx'.format(self.Current_path,self.Current_string)
                    self.WB_Report.save(self.Report_Link_save) 

                #Fill newest report link into "Report_lineEdit"
                self.ui.Report_lineEdit.setText(self.Report_Link_save)

                #-----------------------------------#
                #Progress bar develop later:
                Processing_Satus = 'Extracting Data' #Local variable
                for i in range(10):        
                    self.ui.Status_Console_textBrowser.setStyleSheet("color: rgb(0, 0, 0);") #set text color
                    Processing_Satus += '.'
                    self.ui.Status_Console_textBrowser.setText(Processing_Satus)
                    QtTest.QTest.qWait(500)
                self.ui.Status_Console_textBrowser.setStyleSheet("font-weight: bold; color: green") #set text color
                self.ui.Status_Console_textBrowser.setText('Extracted Done!')
                #-----------------------------------#

    # ANCHOR: Billing Execute
    def Execute_Billing_Function(self):
        try:
            #re-assign
            self.Allowcate_link = self.ui.Allocate_lineEdit.text()
            self.ProjectList_link = self.ui.Projects_lineEdit.text()
            #Check condition:
            if self.Allowcate_link == nullcontext or len(self.Allowcate_link) == 0:
                self.ui.Status_Console_textBrowser.setStyleSheet("color: rgb(255, 0, 0);") #set text color
                self.ui.Status_Console_textBrowser.setText('Please insert Allocate file')
            elif self.ProjectList_link == nullcontext or len(self.ProjectList_link) == 0:
                self.ui.Status_Console_textBrowser.setStyleSheet("color: rgb(255, 0, 0);") #set text color
                self.ui.Status_Console_textBrowser.setText('Please insert Projects list file')
            elif self.ui.Allocate_G4_Sheet_lineEdit.text() == nullcontext or \
                self.ui.Allocate_G5_Sheet_lineEdit.text() == nullcontext or \
                self.ui.Projects_Sheet_lineEdit.text() == nullcontext:
                self.ui.Status_Console_textBrowser.setStyleSheet("color: rgb(255, 0, 0);") #set text color
                self.ui.Status_Console_textBrowser.setText('Please insert sheet name')
            elif self.ui.Allocate_G4_Sheet_lineEdit.text() == self.ui.Allocate_G5_Sheet_lineEdit.text():
                    self.ui.Status_Console_textBrowser.setStyleSheet("color: rgb(255, 0, 0);") #set text color
                    self.ui.Status_Console_textBrowser.setText('Sheet Gen4 and Gen5 should be different')
            elif (self.Format_DateFrom == int(0)) or (self.Format_DateTo == int(0)):
                self.ui.Status_Console_textBrowser.setStyleSheet("color: rgb(255, 0, 0);") #set text color
                self.ui.Status_Console_textBrowser.setText('Please insert date')
            else:
                if self.Year_DateFrom > self.Year_DateTo:
                    self.ui.Status_Console_textBrowser.setStyleSheet("color: rgb(255, 0, 0);")
                    self.ui.Status_Console_textBrowser.setText('INVALID INSERT VALUE: Wrong YEAR value!\nPlease try again')
                elif (self.Year_DateFrom <= self.Year_DateTo) and (self.Month_DateFrom > self.Month_DateTo):
                    self.ui.Status_Console_textBrowser.setStyleSheet("color: rgb(255, 0, 0);") #set text color
                    self.ui.Status_Console_textBrowser.setText('INVALID INSERT VALUE: Wrong MONTH value!\nPlease try again')
                elif (self.Year_DateFrom == self.Year_DateTo) and (self.Month_DateFrom == self.Month_DateTo) and (self.Day_DateFrom > self.Day_DateTo):
                    self.ui.Status_Console_textBrowser.setStyleSheet("color: rgb(255, 0, 0);") #set text color
                    self.ui.Status_Console_textBrowser.setText('INVALID INSERT VALUE: Wrong DAY value!\nPlease try again')
                else: 
                    self.ui.Status_Console_textBrowser.clear()
                    #----------PASSED CONDITION----------#
                    now = datetime.now()
                    dt_string = now.strftime("%d-%m-%Y_%H.%M.%S")

                    #################################
                    #FIXME : ĐIỀN MERGE CODE VÔ ĐÂY:#
                    #################################

                    #####################################
                    ### STAGE 2: Export Sumary report ###
                    #####################################
                    Sum = Template_UI.Summary(self.Report_Link_save)
                    Num = 3
                    LW = load_workbook(self.Report_Link_save)
                    if 'Summary' not in LW.sheetnames:
                        Summary_Sheet = LW.create_sheet('Summary')
                        #Final result:
                        Final_Result_Sheet = LW['Final_Result']
                        sum_headings = ['Project'] + ['Real Hour'] + ['Edit Hour']
                        Summary_Sheet.append(sum_headings)
                        for i in Sum:
                            Summary_Sheet.append([i,int(Sum[i])])
                            Final_Result_Sheet['P{0}'.format(Num)] = i
                            Final_Result_Sheet['Q{0}'.format(Num)] = int(Sum[i])
                            Final_Result_Sheet['R{0}'.format(Num)] = int(Sum[i])/155
                            Num += 1
                    else:
                        Summary_Sheet = LW['Summary']
                        #Final result:
                        Final_Result_Sheet = LW['Final_Result']
                        hour = ''
                        Num = 3
                        temp = []
                        for i in range(1, Summary_Sheet.max_column + 1):
                            char = get_column_letter(i)
                            if Summary_Sheet[char + str(1)].value == "Real Hour":
                                hour = char 
                                break
                        for i in range(2, Summary_Sheet.max_row + 1):
                            Prj = Summary_Sheet["A" + str(i)].value  
                            Summary_Sheet[hour + str(i)].value = Sum[Prj]
                            #Final result:
                            temp.append(Prj)
                        for i in Sum:
                            if i not in temp:
                                Summary_Sheet.append([i,int(Sum[i])])
                                Final_Result_Sheet['P{0}'.format(Num)] = i
                                Final_Result_Sheet['Q{0}'.format(Num)] = int(Sum[i])
                                Final_Result_Sheet['R{0}'.format(Num)] = int(Sum[i])/155
                                Num += 1
                    LW.save(self.Report_Link_save)

                    ####################################
                    ### STAGE 3: Export final result ###
                    ####################################
                    Sum = Template_UI.Get_Hour(self.Report_Link_save)
                    Num = 3
                    LW = load_workbook(self.Report_Link_save)
                    Billing_ws_5 = LW['Gen_5']
                    Billing_ws_4 = LW['Gen_4']
                    ProjectDict_Step_3 = Template_UI.Project_Dict_Step_3(Billing_ws_4,Billing_ws_5)
                    
                    if 'Split_Time' not in LW.sheetnames:
                        ee_splip_time = LW.create_sheet('Split_Time')
                        split_headings = ['Name_Employee'] + ['NumberID'] + ['Project'] + ['BM_Number'] + ['Task_ID'] + ['Hour']
                        ee_splip_time.append(split_headings)
                        S = Template_UI.Split_Time(Sum,EmployeeDict)
                        for i in S:
                            if S[i] == {}:
                                ee_splip_time.append([i,EmployeeDict[i]['NumberID'],"","","",int(155)])
                            for j in S[i]:
                                if j not in ProjectDict_Step_3:
                                    ee_splip_time.append([i,EmployeeDict[i]['NumberID'],j,"","",int(S[i][j])])
                                else:
                                    ee_splip_time.append([i,EmployeeDict[i]['NumberID'],j,ProjectDict_Step_3[j]['BM_Number'],ProjectDict_Step_3[j]['Task_ID'],int(S[i][j])])
                        #Final result:
                        Final_Result_Sheet = LW['Final_Result']
                        for index_employees in EmployeeDict:
                            Final_Result_Sheet['V{0}'.format(Num)] = index_employees
                            Num += 1
                    else:
                        del LW['Split_Time']
                        ee_splip_time = LW.create_sheet('Split_Time')
                        split_headings = ['Name_Employee'] + ['NumberID'] + ['Project'] + ['BM_Number'] + ['Task_ID'] + ['Hour']
                        ee_splip_time.append(split_headings)
                        S = Template_UI.Split_Time(Sum,EmployeeDict)
                        for i in S:
                            if S[i] == {}:
                                ee_splip_time.append([i,EmployeeDict[i]['NumberID'],"","","",int(155)])
                            for j in S[i]:
                                if j not in ProjectDict_Step_3:
                                    ee_splip_time.append([i,EmployeeDict[i]['NumberID'],j,"","",int(S[i][j])])
                                else:
                                    ee_splip_time.append([i,EmployeeDict[i]['NumberID'],j,ProjectDict_Step_3[j]['BM_Number'],ProjectDict_Step_3[j]['Task_ID'],int(S[i][j])])
                    
                        #Final result:
                        Final_Result_Sheet = LW['Final_Result']
                        for index_employees in EmployeeDict:
                            Final_Result_Sheet['V{0}'.format(Num)] = index_employees
                            Num += 1

                    ################################
                    #------------REPORT------------#
                    LW.save(self.Report_Link_save)
                    Template_UI.CALCULATE_155_HOURS(self.Report_Link_save, self.Current_path)
                    ################################
                    #-----------------------------------#
                    #Progress bar develop later:
                    Processing_Satus = 'Processing' #Local variable
                    for i in range(10):        
                        self.ui.Status_Console_textBrowser.setStyleSheet("color: rgb(0, 0, 0);") #set text color
                        Processing_Satus += '.'
                        self.ui.Status_Console_textBrowser.setText(Processing_Satus)
                        QtTest.QTest.qWait(500)
                    self.ui.Status_Console_textBrowser.setStyleSheet("font-weight: bold; color: green") #set text color
                    self.ui.Status_Console_textBrowser.setText('Successfully!')
                    #-----------------------------------#
        except PermissionError:
            self.ui.Status_Console_textBrowser.setStyleSheet("color: rgb(255, 0, 0);") #set text color
            self.ui.Status_Console_textBrowser.setText('Please close report before this action')

    # ANCHOR: Open report:       
    def Open_Report_Function(self):
        if self.ui.Report_lineEdit.text() == '':
            self.ui.Status_Console_textBrowser.setStyleSheet("color: rgb(255, 0, 0);") #set text color
            self.ui.Status_Console_textBrowser.setText('Please insert report file')
        else:
            self.ui.Status_Console_textBrowser.clear()
            Report = self.ui.Report_lineEdit.text()
            excel = gencache.EnsureDispatch('Excel.Application')
            wb = Template_UI.openWorkbook(excel, Report)
            ws = wb.Worksheets('Gen_4') 
            excel.Visible = True    

    def OpenFolderReport_pushButton_Function(self):
        path = '{0}/Report'.format(self.Current_path)
        os.system(f'start {os.path.realpath(path)}')    
    
    # ANCHOR: Re_Calculate_Report_Function
    def Re_Calculate_Report_Function(self):
        #Check condition:
        try:
            self.ui.Status_Console_textBrowser.clear()
            Template_UI.CALCULATE_155_HOURS(self.ui.Report_lineEdit.text(), self.Current_path)
            #-----------------------------------#
            #Progress bar develop later:
            Processing_Satus = 'Processing' #Local variable
            for i in range(10):        
                self.ui.Status_Console_textBrowser.setStyleSheet("color: rgb(0, 0, 0);") #set text color
                Processing_Satus += '.'
                self.ui.Status_Console_textBrowser.setText(Processing_Satus)
                QtTest.QTest.qWait(500)
            self.ui.Status_Console_textBrowser.setStyleSheet("font-weight: bold; color: green") #set text color
            self.ui.Status_Console_textBrowser.setText('Calculated Done!!')
                #-----------------------------------#
        except InvalidFileException:
            self.ui.Status_Console_textBrowser.setStyleSheet("color: rgb(255, 0, 0);") #set text color
            self.ui.Status_Console_textBrowser.setText('Invalid file!')
        except PermissionError:
            self.ui.Status_Console_textBrowser.setStyleSheet("color: rgb(255, 0, 0);") #set text color
            self.ui.Status_Console_textBrowser.setText('Please close report before this action')



    # ANCHOR: Test zone                 
    def DO_SOMETHING_TEST_1(self):
        self.ui.Status_Console_textBrowser.setStyleSheet("font-weight: bold; color: red") #set text color
        self.ui.Status_Console_textBrowser.setText('NOT IN THE TEST MODE')
        
    def DO_SOMETHING_TEST(self):
        self.ui.Status_Console_textBrowser.setStyleSheet("font-weight: bold; color: red") #set text color
        self.ui.Status_Console_textBrowser.setText('NOT IN THE TEST MODE')
            


# ## ------------------------ Helper Fucntion -----------------------------------
    # def LoadInitValue(self):
    #     self.template_init = os.path.join(THIS_FILE, r'lib\template\template_init.json')
    #     usename = os.getlogin()
    #     self.init_value = os.path.join(r'C:\Users', usename, r'AppData\Toolchain.json')

    #     if os.path.isdir(os.path.join(THIS_FILE, 'lib/log')):
    #         pass
    #     else:
    #         os.mkdir(os.path.join(THIS_FILE, 'lib/log'))
    #     f = open('lib/log/DataLog.log', 'a+')
    #     f.truncate(0)
    #     lib.writeLog('Info', 'Load initial config...')

    #     with open(self.template_init, 'r') as json_f:
    #             json_data_template = json.load(json_f)
    #     try:
    #         with open(self.init_value, 'r') as json_f:
    #             json_data = json.load(json_f)
    #     except:
    #         with open(self.init_value, 'w') as json_f:
    #             json.dump(json_data_template, json_f, indent=4)
    #         with open(self.init_value, 'r') as json_f:
    #             json_data = json.load(json_f)

## --------------- Main ----------------------------------
if __name__ == "__main__":
    app = QApplication(sys.argv) 
    main_win = MainWindow()
    main_win.show()
    sys.exit(app.exec())